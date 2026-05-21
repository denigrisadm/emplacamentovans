import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from dateutil.relativedelta import relativedelta
from collections import Counter
import os, re, unicodedata, json, hashlib
from io import BytesIO
import datetime
import base64

st.set_page_config(
    page_title="Comercial De Nigris",
    page_icon="🚚",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ════════════════════════════════════════════════════════════════
# PWA
# ════════════════════════════════════════════════════════════════
_ICON_B64 = ""

def inject_pwa():
    manifest = {
        "name": "Emplacamento Trucks",
        "short_name": "Emp. Trucks",
        "description": "Inteligência Comercial De Nigris",
        "start_url": ".",
        "display": "standalone",
        "background_color": "#0a1628",
        "theme_color": "#0a1628",
        "icons": []
    }
    import json as _json
    manifest_b64 = __import__("base64").b64encode(_json.dumps(manifest).encode()).decode()
    st.markdown(f"""
    <link rel="manifest" href="data:application/manifest+json;base64,{manifest_b64}">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
    <meta name="apple-mobile-web-app-title" content="Emp. Trucks">
    <meta name="theme-color" content="#0a1628">
    <meta name="mobile-web-app-capable" content="yes">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    """, unsafe_allow_html=True)

inject_pwa()

LOGO_B64 = ""

DATA_DIR      = "data"
CARTEIRA_FILE = os.path.join(DATA_DIR, "CARTEIRA.xlsx")
EMP_FILE      = os.path.join(DATA_DIR, "EMPLACAMENTOS.xlsx")
USERS_FILE    = os.path.join(DATA_DIR, "users.json")

NOMES_DENIGRIS = ["COMERCIAL DE VEICULOS DE NIGRIS LTDA"]
REGEX_DENIGRIS = "|".join([re.escape(n) for n in NOMES_DENIGRIS])

def is_denigris(serie):
    return serie.astype(str).str.upper().str.contains(REGEX_DENIGRIS, na=False, regex=True)

MESES_PT = {1:"Janeiro",2:"Fevereiro",3:"Março",4:"Abril",5:"Maio",6:"Junho",
            7:"Julho",8:"Agosto",9:"Setembro",10:"Outubro",11:"Novembro",12:"Dezembro"}

# ════════════════════════════════════════════════════════════════
# CSS — Mobile-first, corporativo azul+dourado
# ════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700;800&display=swap');
*{font-family:'Plus Jakarta Sans',sans-serif!important;box-sizing:border-box;}

[data-testid="stAppViewContainer"]{background:#f4f6fb!important;color:#1a1e2a!important;}
[data-testid="stSidebar"],[data-testid="collapsedControl"]{display:none!important;}
.main .block-container{padding:0!important;max-width:100%!important;}
section[data-testid="stVerticalBlock"]{gap:0!important;}
div[data-testid="stVerticalBlockBorderWrapper"]{padding:0!important;}

/* ── TOPBAR ── */
.topbar{
  background:linear-gradient(135deg,#0a1628 0%,#0d1f3c 100%);
  border-bottom:2px solid #c8a84b;
  padding:0 16px; height:56px;
  display:flex; align-items:center; justify-content:space-between;
  position:sticky; top:0; z-index:9999;
}
.topbar-logo span{font-size:16px;font-weight:800;color:#fff;}
.topbar-user{display:flex;align-items:center;gap:8px;}
.topbar-avatar{
  width:32px;height:32px;
  background:linear-gradient(135deg,#c8a84b,#e8c96b);
  border-radius:50%;
  display:flex;align-items:center;justify-content:center;
  font-size:12px;font-weight:800;color:#0a1628;flex-shrink:0;
}
.topbar-info{text-align:right;}
.topbar-name{font-size:11px;color:#fff;font-weight:700;line-height:1.3;}
.topbar-role{font-size:9px;color:#c8a84b;text-transform:uppercase;letter-spacing:1px;}

/* ── NAV BUTTONS ── */
div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"] > div > div > .stButton > button {
  background:linear-gradient(135deg,#0d1f3c 0%,#1a3355 100%)!important;
  color:#c8a84b!important;
  border:1.5px solid rgba(200,168,75,0.3)!important;
  border-radius:12px!important;
  font-weight:700!important;
  font-size:11px!important;
  padding:10px 4px!important;
  box-shadow:0 3px 10px rgba(10,22,40,0.2)!important;
  transition:all 0.18s ease!important;
  min-height:46px!important;
  letter-spacing:0.1px!important;
}
div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"] > div > div > .stButton > button:hover {
  background:linear-gradient(135deg,#c8a84b 0%,#e8c96b 100%)!important;
  color:#0a1628!important;
  border-color:#c8a84b!important;
  box-shadow:0 6px 18px rgba(200,168,75,0.3)!important;
  transform:translateY(-2px)!important;
}
div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"]:last-child > div > div > .stButton > button {
  background:linear-gradient(135deg,#3a0a0a 0%,#5a1010 100%)!important;
  color:#ff8870!important;
  border-color:rgba(255,100,80,0.35)!important;
}
div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"]:last-child > div > div > .stButton > button:hover {
  background:linear-gradient(135deg,#cc3300 0%,#ee4411 100%)!important;
  color:#fff!important;
  border-color:#cc3300!important;
}
div[data-testid="stHorizontalBlock"]:has(.stButton){
  background:rgba(244,246,251,0.95)!important;
  padding:8px 12px 6px 12px!important;
  border-bottom:1px solid rgba(200,168,75,0.12)!important;
  position:sticky!important;
  top:56px!important;
  z-index:9990!important;
  backdrop-filter:blur(12px)!important;
  -webkit-backdrop-filter:blur(12px)!important;
}

/* ── MOBILE ── */
@media (max-width: 640px) {
  .page-wrap{padding:10px 8px 80px 8px!important;}
  .kpi-grid{grid-template-columns:repeat(2,1fr)!important;gap:7px!important;}
  .kpi-grid-3{grid-template-columns:repeat(2,1fr)!important;}
  .kpi-value{font-size:20px!important;}
  .topbar{padding:0 10px!important;height:50px!important;}
  .topbar-name{font-size:10px!important;}
  .hero-name{font-size:14px!important;}
  .hero-stats{flex-wrap:wrap!important;}
  .hero-stat{min-width:70px!important;}
  .info-label{width:80px!important;min-width:80px!important;}
  div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"] > div > div > .stButton > button {
    font-size:9px!important;
    padding:7px 1px!important;
    min-height:40px!important;
    border-radius:8px!important;
  }
  div[data-testid="stHorizontalBlock"]:has(.stButton){
    top:50px!important;
    padding:6px 8px 5px 8px!important;
  }
}

/* ── CONTEÚDO ── */
.page-wrap{padding:16px 14px 80px 14px;max-width:920px;margin:0 auto;}
.page-header{margin-bottom:16px;}
.page-header h1{font-size:20px;font-weight:800;color:#0a1628;margin:0 0 2px 0;}
.page-header p{font-size:11px;color:#8a95b0;margin:0;}

/* ── LOGIN ── */
.login-bg{
  min-height:100vh;
  background:linear-gradient(160deg,#0a1628 0%,#0d1f3c 60%,#0a1628 100%);
  display:flex; align-items:center; justify-content:center;
  padding:24px 16px;
}

/* ── KPI CARDS ── */
.kpi-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:10px;margin-bottom:16px;}
.kpi-grid-3{display:grid;grid-template-columns:repeat(3,1fr);gap:9px;margin-bottom:16px;}
.kpi-card{
  background:#fff; border-radius:14px;
  padding:14px 12px; position:relative; overflow:hidden;
  box-shadow:0 2px 10px rgba(0,0,0,0.05);
  border:1px solid #f0f2f8;
}
.kpi-card::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;background:linear-gradient(90deg,#0a1628,#c8a84b);}
.kpi-label{font-size:9px;text-transform:uppercase;letter-spacing:1.2px;color:#8a95b0;margin-bottom:5px;font-weight:700;}
.kpi-value{font-size:24px;font-weight:800;color:#0a1628;line-height:1;}
.kpi-value.green{color:#007030;}
.kpi-value.red{color:#a02020;}
.kpi-value.blue{color:#0044aa;}
.kpi-value.gold{color:#b8860b;}
.kpi-sub{font-size:10px;color:#8a95b0;margin-top:3px;}

/* ── SECTION TITLE ── */
.sec-title{
  font-size:13px;font-weight:800;color:#0a1628;
  margin:20px 0 10px 0;padding-bottom:7px;
  border-bottom:2px solid #e8ecf5;
  display:flex;align-items:center;gap:5px;
}

/* ── CARDS QUADRANTES ── */
.quadrant{
  background:#fff;border-radius:14px;padding:14px;
  box-shadow:0 2px 10px rgba(0,0,0,0.05);
  border:1px solid #f0f2f8;
  margin-bottom:12px;
}
.quadrant-header{display:flex;align-items:center;justify-content:space-between;margin-bottom:10px;}
.quadrant-title{font-size:12px;font-weight:800;color:#0a1628;}
.quadrant-count{
  min-width:42px;height:42px;border-radius:10px;
  display:flex;align-items:center;justify-content:center;
  font-size:18px;font-weight:800;
}
.q-red{background:#ffeee8;color:#cc3300;}
.q-yellow{background:#fff8e0;color:#b8860b;}
.q-green{background:#e8f8ee;color:#007030;}
.q-blue{background:#e8f0ff;color:#0044aa;}

/* ── INFO TABLE ── */
.info-card{background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.04);border:1px solid #f0f2f8;margin-bottom:12px;}
.info-row{display:flex;align-items:flex-start;padding:9px 12px;border-bottom:1px solid #f4f6fb;}
.info-row:last-child{border-bottom:none;}
.info-label{font-size:9px;text-transform:uppercase;letter-spacing:1px;color:#8a95b0;width:110px;min-width:110px;padding-top:2px;font-weight:700;}
.info-value{font-size:12px;color:#1a2a3a;font-weight:500;flex:1;word-break:break-word;}
.info-value strong{color:#0a1628;font-weight:800;}

/* ── CONTACT BUTTONS ── */
.contact-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:7px;margin-bottom:12px;}
.contact-btn{
  display:flex;flex-direction:column;align-items:center;justify-content:center;
  padding:10px 6px;border-radius:12px;text-decoration:none!important;
  border:1px solid; gap:3px; transition:all 0.2s;
  font-size:10px;font-weight:700;text-align:center;
}
.btn-phone{background:#e8f4ff;color:#0055bb;border-color:#c0d8f8;}
.btn-whatsapp{background:#e8fff2;color:#007a3d;border-color:#a0e8c0;}
.btn-email{background:#fff8e8;color:#a06000;border-color:#f0d090;}

/* ── BADGES ── */
.badge{display:inline-flex;align-items:center;padding:3px 9px;border-radius:20px;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.5px;gap:3px;}
.badge-top{background:#fff8e0;color:#8a6000;border:1px solid #f0d080;}
.badge-ativo{background:#e0f8e8;color:#007030;border:1px solid #80d090;}
.badge-baixa{background:#fff4e0;color:#a05000;border:1px solid #f0c070;}
.badge-inativo{background:#ffeee8;color:#a02020;border:1px solid #f0a090;}

/* ── ALERTS ── */
.alert-red{background:#ffeee8;border-left:4px solid #cc3300;border-radius:8px;padding:10px 12px;color:#8a2200;margin:8px 0;font-size:12px;}
.alert-yellow{background:#fff8e0;border-left:4px solid #e8a000;border-radius:8px;padding:10px 12px;color:#8a5000;margin:8px 0;font-size:12px;}
.alert-green{background:#e8f8ee;border-left:4px solid #00aa55;border-radius:8px;padding:10px 12px;color:#005530;margin:8px 0;font-size:12px;}
.alert-blue{background:#e8f0ff;border-left:4px solid #0055cc;border-radius:8px;padding:10px 12px;color:#003080;margin:8px 0;font-size:12px;}
.alert-gray{background:#f4f6fb;border-left:4px solid #a0a8b8;border-radius:8px;padding:10px 12px;color:#4a5568;margin:8px 0;font-size:12px;}

/* ── SOCIO CARD ── */
.socio-card{background:#f8f9fc;border:1px solid #e8ecf5;border-radius:12px;padding:12px;margin-bottom:8px;}
.socio-name{font-size:13px;font-weight:800;color:#0a1628;margin-bottom:2px;}
.socio-role{font-size:10px;color:#0055aa;text-transform:uppercase;letter-spacing:0.8px;font-weight:700;}

/* ── CLIENT HERO ── */
.client-hero{
  background:linear-gradient(135deg,#0a1628,#0d1f3c);
  border-radius:16px;padding:18px;margin-bottom:14px;
  position:relative;overflow:hidden;
}
.client-hero::after{content:'';position:absolute;right:-40px;top:-40px;
  width:160px;height:160px;background:radial-gradient(circle,rgba(200,168,75,0.08) 0%,transparent 70%);border-radius:50%;}
.hero-name{font-size:16px;font-weight:800;color:#fff;line-height:1.3;margin-bottom:3px;}
.hero-cnpj{font-size:11px;color:#c8a84b;font-family:monospace;letter-spacing:0.8px;}
.hero-stats{display:flex;gap:8px;margin-top:12px;}
.hero-stat{background:rgba(255,255,255,0.07);border-radius:10px;padding:9px 12px;text-align:center;flex:1;}
.hero-stat-val{font-size:20px;font-weight:800;color:#c8a84b;}
.hero-stat-lbl{font-size:8px;color:#8899bb;text-transform:uppercase;letter-spacing:0.8px;}

/* ── UPLOAD SECTION ── */
.upload-box{background:#fff;border:2px dashed #d0d8ee;border-radius:12px;padding:14px;margin-bottom:10px;}
.upload-title{font-size:11px;font-weight:800;text-transform:uppercase;letter-spacing:0.8px;color:#0a1628;margin-bottom:5px;}

/* ── STREAMLIT OVERRIDES ── */
.stTextInput input{background:#f8f9fc!important;border:1.5px solid #dde2f0!important;color:#1a1e2a!important;border-radius:10px!important;font-size:14px!important;padding:11px 13px!important;}
.stTextInput input:focus{border-color:#0a1628!important;box-shadow:0 0 0 3px rgba(10,22,40,0.07)!important;}
.stButton>button{
  background:linear-gradient(135deg,#0a1628,#142038)!important;
  color:#c8a84b!important;border:none!important;border-radius:10px!important;
  font-weight:700!important;padding:11px 18px!important;font-size:13px!important;
  letter-spacing:0.2px!important;width:100%;
  box-shadow:0 3px 12px rgba(10,22,40,0.18)!important;
  transition:all 0.18s!important;
}
.stButton>button:hover{transform:translateY(-1px)!important;box-shadow:0 5px 16px rgba(10,22,40,0.25)!important;}
.stButton>button[kind="secondary"]{background:#f0f2f7!important;color:#0a1628!important;border:1.5px solid #dde2f0!important;box-shadow:none!important;}
.stSelectbox>div>div,.stMultiSelect>div>div{background:#f8f9fc!important;border:1.5px solid #dde2f0!important;border-radius:10px!important;}
.stDataFrame{border-radius:10px!important;overflow:hidden!important;}
[data-testid="stDataFrame"] thead tr th{background:#0a1628!important;color:#c8a84b!important;font-weight:700!important;font-size:10px!important;text-transform:uppercase!important;}
.stTabs [data-baseweb="tab-list"]{background:#f0f2f7;border-radius:10px;padding:3px;gap:3px;}
.stTabs [data-baseweb="tab"]{background:transparent;border-radius:8px;font-weight:700;font-size:11px;color:#4a5568;padding:7px 12px!important;}
.stTabs [aria-selected="true"]{background:#0a1628!important;color:#c8a84b!important;}
.stDownloadButton>button{background:linear-gradient(135deg,#0a1628,#142038)!important;color:#c8a84b!important;font-weight:700!important;border:none!important;border-radius:10px!important;}
div[data-testid="stExpander"]{background:#fff;border:1.5px solid #e8ecf5!important;border-radius:12px!important;}

/* ── RECEITA FEDERAL CARD ── */
.rf-card{background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:16px;margin-bottom:14px;box-shadow:0 2px 8px rgba(0,0,0,0.05);}
.rf-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:10px;font-size:12px;margin-bottom:10px;}
@media(max-width:640px){.rf-grid{grid-template-columns:1fr!important;}}
.rf-item{background:#f8fafc;padding:9px;border-radius:8px;}
.rf-item-label{color:#8a95b0;font-size:9px;margin-bottom:2px;text-transform:uppercase;letter-spacing:0.8px;}
.rf-item-value{font-weight:600;color:#0a1628;font-size:12px;}

/* Hide streamlit extras */
#MainMenu,footer,[data-testid="stToolbar"]{display:none!important;}
[data-testid="stForm"]{border:none!important;padding:0!important;background:transparent!important;box-shadow:none!important;}
</style>
""", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════
# HELPERS
# ════════════════════════════════════════════════════════════════
def norm_str(s):
    if not isinstance(s, str): s = str(s)
    s = s.upper().strip()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', s)

def norm_str_series(series):
    s = series.fillna('').astype(str).str.upper().str.strip()
    s = s.apply(lambda x: unicodedata.normalize('NFKD', x))
    s = s.str.encode('ascii', errors='ignore').str.decode('ascii')
    return s.str.replace(r'\s+', ' ', regex=True).str.strip()

def norm_cnpj(s):
    return re.sub(r"[.\-/]", "", str(s)).strip()

def norm_cep(v):
    if pd.isna(v): return ""
    try:
        n = int(float(v))
        return str(n).zfill(8)[:8]
    except (ValueError, TypeError):
        s = re.sub(r"\D", "", str(v))
        return s.zfill(8)[:8] if s else ""

def safe_str(v, fallback="—"):
    s = str(v).strip() if not pd.isna(v) else ""
    return fallback if s in ["nan","None","NaN","","NAN"] else s

def safe_atividade(d):
    s = safe_str(d.get("ATIVIDADE_ECONOMICA",""))
    return (s[:90]+"...") if len(s) > 90 else s

def format_tel(ddd, num):
    try:
        d = str(int(float(ddd))) if pd.notna(ddd) else ""
        n = str(int(float(num))) if pd.notna(num) else ""
        return f"({d}) {n}" if d and n else ""
    except: return ""

def make_fone_num(ddd, num):
    try:
        d = str(int(float(ddd))) if pd.notna(ddd) else ""
        n = str(int(float(num))) if pd.notna(num) else ""
        return f"55{d}{n}" if d and n else ""
    except: return ""

def hash_senha(s):
    return hashlib.sha256(s.encode()).hexdigest()

def get_modes(series, top=3):
    s = series.dropna().astype(str)
    s = s[~s.isin(["nan","None","NaN","","N/A"])]
    return [x for x,_ in Counter(s).most_common(top)] if not s.empty else []

def calc_prediction(dates):
    """Retorna (label_str, predicted_date) ou (None, None). Robusto a erros."""
    try:
        if not dates or len(dates) < 2:
            return None, None
        clean = []
        for d in dates:
            try:
                ts = pd.Timestamp(d)
                if pd.notna(ts):
                    clean.append(ts)
            except Exception:
                continue
        if len(clean) < 2:
            return None, None
        clean = sorted(clean)
        intervals = []
        for i in range(1, len(clean)):
            try:
                rd = relativedelta(clean[i], clean[i-1])
                m = rd.years * 12 + rd.months
                if m > 0:
                    intervals.append(m)
            except Exception:
                continue
        if not intervals:
            return None, None
        avg = max(sum(intervals) / len(intervals), 1)
        predicted = clean[-1] + relativedelta(months=int(round(avg)))
        return f"{MESES_PT[predicted.month]} de {predicted.year}", predicted
    except Exception:
        return None, None

def badge_class(cls):
    c = str(cls).lower()
    if "top" in c: return "badge-top"
    if "baixa" in c: return "badge-baixa"
    if "ativo" in c: return "badge-ativo"
    return "badge-inativo"

def logo_img(height=32):
    if LOGO_B64:
        return f'<img src="data:image/png;base64,{LOGO_B64}" style="height:{height}px;object-fit:contain;">'
    return f'<span style="font-size:{max(14,height//2)}px;font-weight:800;color:#fff;">Comercial De Nigris</span>'

# ════════════════════════════════════════════════════════════════
# USUÁRIOS — GitHub API
# ════════════════════════════════════════════════════════════════
def _gh_secrets():
    try:
        token  = st.secrets.get("GH_TOKEN","")
        repo   = st.secrets.get("GH_REPO","") or "denigrisadm/emplacamentovans"
        branch = st.secrets.get("GH_BRANCH","main")
        return token, repo, branch
    except Exception:
        return "", "denigrisadm/emplacamentovans", "main"

def _gh_get_file(api_url, token):
    import urllib.request
    try:
        req = urllib.request.Request(
            api_url,
            headers={"Authorization": f"token {token}",
                     "Accept": "application/vnd.github.v3+json"}
        )
        with urllib.request.urlopen(req, timeout=12) as r:
            data = json.loads(r.read().decode())
        content = base64.b64decode(data["content"].replace("\n",""))
        return content, data.get("sha","")
    except Exception:
        return None, ""

def _gh_put_file(api_url, token, branch, content_bytes, sha, message="update users.json"):
    import urllib.request
    try:
        payload = json.dumps({
            "message": message,
            "content": base64.b64encode(content_bytes).decode(),
            "sha": sha,
            "branch": branch
        }).encode()
        req = urllib.request.Request(
            api_url, data=payload, method="PUT",
            headers={"Authorization": f"token {token}",
                     "Content-Type": "application/json",
                     "Accept": "application/vnd.github.v3+json"}
        )
        with urllib.request.urlopen(req, timeout=15): pass
        return True
    except Exception:
        return False

def load_users():
    token, repo, branch = _gh_secrets()
    if token and repo:
        api_url = f"https://api.github.com/repos/{repo}/contents/data/users.json"
        content, _ = _gh_get_file(api_url, token)
        if content:
            try:
                users = json.loads(content.decode("utf-8"))
                os.makedirs(DATA_DIR, exist_ok=True)
                with open(USERS_FILE,"w",encoding="utf-8") as f:
                    f.write(json.dumps(users, ensure_ascii=False, indent=2))
                return users
            except Exception:
                pass
    if os.path.exists(USERS_FILE):
        try:
            with open(USERS_FILE,"r",encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {
        "ADMIN": {
            "senha_hash": hash_senha("admin2025"),
            "perfil": "gestor",
            "nome": "Administrador",
            "ultimo_acesso": None
        }
    }

def save_users(users):
    os.makedirs(DATA_DIR, exist_ok=True)
    content_str = json.dumps(users, ensure_ascii=False, indent=2)
    content_bytes = content_str.encode("utf-8")
    gh_ok = False
    gh_err = "GitHub não configurado"
    token, repo, branch = _gh_secrets()
    if token and repo:
        try:
            api_url = f"https://api.github.com/repos/{repo}/contents/data/users.json"
            _, sha = _gh_get_file(api_url, token)
            resultado = _gh_put_file(api_url, token, branch, content_bytes, sha or "")
            gh_ok = bool(resultado)
            gh_err = "" if gh_ok else "Falha no envio"
        except Exception as e:
            gh_ok = False
            gh_err = str(e)
    try:
        with open(USERS_FILE, "w", encoding="utf-8") as f:
            f.write(content_str)
    except Exception:
        pass
    return gh_ok, gh_err

def registrar_acesso(login):
    users = st.session_state.get("users_db", load_users())
    if login in users:
        try:
            agora_brt = datetime.datetime.utcnow() - datetime.timedelta(hours=3)
            users[login]["ultimo_acesso"] = agora_brt.strftime("%Y-%m-%dT%H:%M:%S")
            save_users(users)
            st.session_state.users_db = users
        except Exception:
            pass

# ════════════════════════════════════════════════════════════════
# CARREGAMENTO DE DADOS
# ════════════════════════════════════════════════════════════════
def load_carteira(src):
    if isinstance(src, BytesIO): src.seek(0)
    df = pd.read_excel(src)
    df.columns = [c.strip() for c in df.columns]
    df["CPF/CNPJ"]  = df["CPF/CNPJ"].astype(str).str.strip()
    df["CNPJ_NORM"] = df["CPF/CNPJ"].astype(str).str.replace(r"\D","",regex=True)
    df["VENDEDOR"]  = df.get("VENDEDOR", pd.Series(dtype=str)).astype(str).str.strip()
    if "CEP" in df.columns:
        df["CEP_norm"] = df["CEP"].apply(norm_cep)
    return df

def load_emplacamentos(src, label=""):
    if isinstance(src, BytesIO): src.seek(0)
    raw = pd.read_excel(src, header=None, nrows=15)
    header_row = 0
    for i in range(len(raw)):
        if any("Chassi" in str(v) for v in raw.iloc[i].tolist()):
            header_row = i; break
    if isinstance(src, BytesIO): src.seek(0)
    df = pd.read_excel(src, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]
    for col in ["CPFCNPJPROPRIETARIO","NOMEPROPRIETARIO","NO_CIDADE","NO_BAIRRO",
                "Placa","Chassi","Concessionário","Modelo","Marca","Segmento"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    df["CNPJ_NORM"] = df["CPFCNPJPROPRIETARIO"].astype(str).str.replace(r"\D", "", regex=True)
    df["Data emplacamento"] = pd.to_datetime(df["Data emplacamento"], dayfirst=True, errors="coerce")
    if df["Data emplacamento"].isna().sum() > len(df) * 0.5:
        df["Data emplacamento"] = pd.to_datetime(df["Data emplacamento"], errors="coerce")
    df["Ano"] = df["Data emplacamento"].dt.year
    df["Mes"] = df["Data emplacamento"].dt.month
    df["Placa_norm"] = df["Placa"].str.replace("-","").str.replace(" ","").str.upper()
    df["NO_CIDADE_NORM"] = norm_str_series(df["NO_CIDADE"])
    df["NO_BAIRRO_NORM"] = norm_str_series(df["NO_BAIRRO"].fillna(""))
    if "NU_CEP" in df.columns:
        df["CEP_norm"] = (df["NU_CEP"].astype(str)
                          .str.replace(r"\D","",regex=True)
                          .str.zfill(8).str[:8])
    else:
        df["CEP_norm"] = ""
    df.dropna(subset=["Ano"], inplace=True)
    df["Ano"] = df["Ano"].astype(int)
    df["Mes"] = df["Mes"].astype(int)
    df["_fonte"] = label
    return df

def merge_emp(dfs):
    m = pd.concat(dfs, ignore_index=True)
    m.drop_duplicates(subset=["Chassi"], keep="last", inplace=True)
    m.sort_values("Data emplacamento", inplace=True)
    return m

def get_vendedores_ativos():
    users = st.session_state.get("users_db", {})
    if not users: return []
    vendedores = [u["nome"] for u in users.values() if u.get("perfil") == "vendedor"]
    return sorted(list(set(vendedores)))

def get_consultor_distribuido(cnpj_norm, vendedores_ativos):
    """Distribui CNPJ de forma determinística e equilibrada — mesmo CNPJ → sempre mesmo vendedor."""
    if not vendedores_ativos:
        return None
    h = int(hashlib.md5(str(cnpj_norm).encode()).hexdigest(), 16)
    idx = h % len(vendedores_ativos)
    return vendedores_ativos[idx]

def get_vendedor_final(cnpj_norm, cart_row, vendedores_ativos):
    """
    Regra de atribuição:
    1. Se está na carteira com vendedor definido → usa ele.
    2. Se não → distribui deterministicamente entre todos os vendedores ativos.
    """
    if cart_row:
        v = safe_str(cart_row.get("VENDEDOR", ""), "").strip()
        if v and v != "—":
            return v
    return get_consultor_distribuido(cnpj_norm, vendedores_ativos)

def load_excel_from_github(filename):
    import urllib.request, urllib.parse
    token, repo, branch = _gh_secrets()
    if not repo:
        local_path = os.path.join("data", filename)
        if os.path.exists(local_path):
            try:
                with open(local_path, "rb") as f:
                    return BytesIO(f.read()), None
            except Exception as e:
                return None, f"Erro lendo local: {e}"
        return None, "Repo não configurado e arquivo local não encontrado"
    full_repo = repo if "/" in repo else f"denigrisadm/{repo}"
    filename_enc = urllib.parse.quote(filename)
    api_url = f"https://api.github.com/repos/{full_repo}/contents/data/{filename_enc}?ref={branch}"
    try:
        req = urllib.request.Request(api_url)
        req.add_header("Accept", "application/vnd.github.v3.raw")
        if token:
            req.add_header("Authorization", f"token {token}")
        with urllib.request.urlopen(req, timeout=30) as r:
            return BytesIO(r.read()), None
    except Exception as e:
        try:
            url_raw = f"https://raw.githubusercontent.com/{full_repo}/{branch}/data/{filename_enc}"
            req2 = urllib.request.Request(url_raw)
            if token:
                req2.add_header("Authorization", f"token {token}")
            with urllib.request.urlopen(req2, timeout=30) as r2:
                return BytesIO(r2.read()), None
        except Exception as e2:
            local_path = os.path.join("data", filename)
            if os.path.exists(local_path):
                try:
                    with open(local_path, "rb") as f:
                        return BytesIO(f.read()), None
                except: pass
            return None, f"API: {e} | RAW: {e2}"

# ════════════════════════════════════════════════════════════════
# RECEITA FEDERAL — consulta robusta com múltiplas APIs
# ════════════════════════════════════════════════════════════════
def consultar_receita_federal(cnpj14):
    """
    Tenta múltiplas APIs públicas em sequência.
    Retorna (dados_dict, erro_str).
    dados_dict = None se todas falharem.
    """
    import urllib.request as _ur
    import urllib.error as _ue
    import ssl

    cnpj_digits = re.sub(r"\D", "", str(cnpj14))
    if len(cnpj_digits) != 14:
        return None, "CNPJ deve ter 14 dígitos"

    # Criar contexto SSL que aceita certificados (algumas APIs têm cert issues)
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    apis = [
        f"https://publica.cnpj.ws/cnpj/{cnpj_digits}",
        f"https://brasilapi.com.br/api/cnpj/v1/{cnpj_digits}",
        f"https://receitaws.com.br/v1/cnpj/{cnpj_digits}",
        f"https://minhareceita.org/{cnpj_digits}",
    ]

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/json",
        "Accept-Language": "pt-BR,pt;q=0.9",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
    }

    last_err = ""
    for url in apis:
        try:
            req = _ur.Request(url, headers=headers)
            with _ur.urlopen(req, timeout=15, context=ctx) as r:
                raw = r.read().decode("utf-8", errors="replace")
                data = json.loads(raw)
                # Verificar se retornou erro da própria API
                if isinstance(data, dict):
                    msg = data.get("message","") or data.get("erro","") or data.get("status","")
                    if str(msg).upper() in ["TRUE","ERROR","CNPJ INVÁLIDO"]:
                        last_err = f"API retornou erro: {msg}"
                        continue
                return data, None
        except _ue.HTTPError as e:
            last_err = f"{url.split('/')[2]}: HTTP {e.code}"
            if e.code == 429:
                import time; time.sleep(0.5)  # rate limit — aguarda um pouco
            continue
        except Exception as e:
            last_err = f"{url.split('/')[2]}: {str(e)[:80]}"
            continue

    return None, f"Todas as APIs falharam. Último erro: {last_err}"

def parse_receita_federal(d):
    """Normaliza resposta de qualquer API para dict padronizado."""
    est = d.get("estabelecimento", d)
    def sv(v, fb="—"):
        if isinstance(v, dict):
            return v.get("descricao") or v.get("nome") or fb
        s = str(v).strip() if v is not None else ""
        return fb if s in ["None","nan","","null","NULL"] else s

    razao    = sv(d.get("razao_social") or d.get("nome",""))
    fantasia = sv(est.get("nome_fantasia") or d.get("fantasia",""))
    situacao = sv(est.get("situacao_cadastral") or d.get("situacao",""))
    abertura = sv(est.get("data_inicio_atividade") or d.get("abertura",""))
    capital  = sv(d.get("capital_social",""))
    porte    = sv(d.get("porte",""))
    nat_jur  = sv(d.get("natureza_juridica",""))

    # Atividade principal
    atv_raw = est.get("atividade_principal") or d.get("atividade_principal",{})
    if isinstance(atv_raw, dict):
        atividade = f"{sv(atv_raw.get('subclasse',''))} — {sv(atv_raw.get('descricao',''))}"
    elif isinstance(atv_raw, list) and atv_raw:
        a0 = atv_raw[0]
        atividade = a0.get("text") or a0.get("descricao","—")
    else:
        atividade = "—"

    # Endereço
    end_parts = [sv(est.get(k,""),"") for k in ["tipo_logradouro","logradouro","numero","complemento","bairro"]]
    end_rf = " ".join(p for p in end_parts if p and p != "—").strip() or "—"
    # Fallback para formato receitaws
    if end_rf == "—":
        end_rf = sv(d.get("logradouro","")) + " " + sv(d.get("numero","")) + " " + sv(d.get("complemento",""))
        end_rf = end_rf.strip() or "—"

    cidade_rf = sv(est.get("cidade",{}) or est.get("municipio","") or d.get("municipio",""))
    uf_rf     = sv(est.get("estado",{}) or est.get("uf","") or d.get("uf",""))
    cep_rf    = sv(est.get("cep","") or d.get("cep",""))
    socios    = d.get("socios") or d.get("qsa") or est.get("socios",[]) or []

    # Formatar capital
    try:
        cap_num = float(str(capital).replace(",",".").replace(" ",""))
        capital_fmt = f"R$ {cap_num:,.2f}".replace(",","X").replace(".",",").replace("X",".")
    except Exception:
        capital_fmt = capital if capital != "—" else "—"

    return dict(
        razao=razao, fantasia=fantasia, situacao=situacao, abertura=abertura,
        capital=capital_fmt, porte=porte, nat_jur=nat_jur, atividade=atividade,
        end_rf=end_rf, cidade_rf=cidade_rf, uf_rf=uf_rf, cep_rf=cep_rf, socios=socios
    )

# ════════════════════════════════════════════════════════════════
# RELATÓRIO XLSX ESTILIZADO
# ════════════════════════════════════════════════════════════════
def gerar_relatorio_emplacamento(emp_mes, emp_area, cnpjs_carteira_set, todos_cnpjs_cart,
                                  sel_mes_lbl, sel_ano, consultor_nome):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"{sel_mes_lbl[:3]} {sel_ano}"

    PRETO = "FF1A1A1A"; BRANCO = "FFFFFFFF"; VERDE_CL = "FF28A745"
    VERMELHO = "FFC0392B"; AZUL = "FF1A3F6F"; CINZA_CL = "FFF2F2F2"
    AMARELO = "FFFFF2CC"; ROSA_CL = "FFFCE8E8"; AZUL_CL = "FFE8F0FF"

    thin = Side(style="thin", color="FFCCCCCC")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(c): return PatternFill("solid", start_color=c, end_color=c)
    def font(bold=False, color=BRANCO, size=11):
        return Font(name="Arial", bold=bold, color=color, size=size)
    def aln(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    for col, w in zip("ABCDEFG", [14,22,40,28,22,35,18]):
        ws.column_dimensions[col].width = w

    # Linha 1 — Título
    ws.merge_cells("A1:G1"); ws.row_dimensions[1].height = 34
    c = ws["A1"]
    c.value = f"EMPLACAMENTO ACUMULADO — {sel_mes_lbl.upper()} DE {sel_ano}"
    c.font = Font(name="Arial", bold=True, color=BRANCO, size=13)
    c.fill = fill(PRETO); c.alignment = aln()

    ws.merge_cells("A2:G2"); ws.row_dimensions[2].height = 16
    c = ws["A2"]
    c.value = f"Consultor: {consultor_nome.title()}"
    c.font = Font(name="Arial", color="FFAAAAAA", size=9, italic=True)
    c.fill = fill(PRETO); c.alignment = aln()

    g1 = emp_mes[~emp_mes["CNPJ_NORM"].isin(todos_cnpjs_cart)]
    g2 = emp_mes[emp_mes["CNPJ_NORM"].isin(cnpjs_carteira_set) & ~is_denigris(emp_mes["Concessionário"])]
    g3 = emp_mes[is_denigris(emp_mes["Concessionário"])]
    total = len(emp_mes)
    def pct(n): return f"{round(n/total*100) if total else 0}%"

    for r_idx in [3,4,5]:
        ws.row_dimensions[r_idx].height = 26 if r_idx == 3 else 22

    # Grupos cabeçalho
    for cols, lbl, cor, val in [
        ("A","EMPLACAMENTOS", "FF1E7E34", total),
        ("BC","NÃO ESTÃO NA CARTEIRA", PRETO, len(g1)),
        ("DE","CARTEIRA → CONCORRÊNCIA", VERMELHO, len(g2)),
        ("FG","COMPROU NA DE NIGRIS", AZUL, len(g3)),
    ]:
        if len(cols) == 1:
            c3 = ws[f"A3"]; c4 = ws[f"A4"]; c5 = ws[f"A5"]
        else:
            c1,c2 = cols[0]+"3", cols[1]+"3"
            ws.merge_cells(f"{c1}:{cols[1]}3")
            ws.merge_cells(f"{cols[0]}4:{cols[1]}4")
            ws.merge_cells(f"{cols[0]}5:{cols[1]}5")
            c3 = ws[f"{cols[0]}3"]; c4 = ws[f"{cols[0]}4"]; c5 = ws[f"{cols[0]}5"]
        for c_obj, v, s in [(c3, lbl, 9), (c4, val, 13), (c5, pct(val) if lbl != "EMPLACAMENTOS" else "100%", 11)]:
            c_obj.value = v; c_obj.fill = fill(cor)
            c_obj.font = Font(name="Arial", bold=True, color=BRANCO, size=s)
            c_obj.alignment = aln(wrap=(s==9)); c_obj.border = borda

    # Linha 6
    ws.row_dimensions[6].height = 20
    ws.merge_cells("A6:C6"); ws["A6"].value = f"CLIENTES ÚNICOS: {emp_mes['CNPJ_NORM'].nunique()}"
    ws["A6"].font = font(True, "FF1A1A1A", 9); ws["A6"].fill = fill("FFE0E0E0"); ws["A6"].alignment = aln("left"); ws["A6"].border = borda
    ni = g3["CNPJ_NORM"].nunique(); tu = emp_mes["CNPJ_NORM"].nunique()
    ws.merge_cells("D6:E6"); ws["D6"].value = f"De Nigris: {ni} ({round(ni/tu*100) if tu else 0}%)"
    ws["D6"].font = font(True, BRANCO, 9); ws["D6"].fill = fill(AZUL); ws["D6"].alignment = aln(); ws["D6"].border = borda
    ws.merge_cells("F6:G6"); ws["F6"].fill = fill("FFE0E0E0"); ws["F6"].border = borda

    # Linha 7 — Top 3
    ws.row_dimensions[7].height = 18; ws.merge_cells("A7:G7")
    ws["A7"].value = "🏆  TOP 3 CLIENTES DO PERÍODO"
    ws["A7"].font = Font(name="Arial", bold=True, color=BRANCO, size=9)
    ws["A7"].fill = fill("FF2C3E50"); ws["A7"].alignment = aln("left"); ws["A7"].border = borda

    top3 = emp_mes.groupby(["CNPJ_NORM","NOMEPROPRIETARIO","NO_CIDADE"]).agg(
        Total=("Chassi","count"), Nigris=("Concessionário", lambda x: is_denigris(x).sum())
    ).reset_index().sort_values("Total", ascending=False).head(3).reset_index(drop=True)
    medals = ["🥇","🥈","🥉"]
    for ti, trow in top3.iterrows():
        r = 8 + ti; ws.row_dimensions[r].height = 18
        pct_n = round(trow["Nigris"]/trow["Total"]*100) if trow["Total"] > 0 else 0
        rf = fill(CINZA_CL if ti%2==0 else BRANCO)
        ws.merge_cells(f"A{r}:B{r}")
        ws[f"A{r}"].value = f"{medals[ti]}  {str(trow['NOMEPROPRIETARIO'])[:38]}"
        ws[f"A{r}"].font = Font(name="Arial", bold=True, color="FF1A1A1A", size=9); ws[f"A{r}"].fill = rf
        ws[f"A{r}"].alignment = aln("left"); ws[f"A{r}"].border = borda
        ws.merge_cells(f"C{r}:D{r}")
        ws[f"C{r}"].value = f"📍 {safe_str(trow['NO_CIDADE'])}"; ws[f"C{r}"].font = Font(name="Arial",size=9,color="FF555555")
        ws[f"C{r}"].fill = rf; ws[f"C{r}"].alignment = aln("left"); ws[f"C{r}"].border = borda
        ws.merge_cells(f"E{r}:F{r}")
        cor_t = "FF1E7E34" if pct_n >= 50 else "FFC0392B"
        ws[f"E{r}"].value = f"{int(trow['Total'])} veículos — {pct_n}% De Nigris"
        ws[f"E{r}"].font = Font(name="Arial", bold=True, size=9, color=cor_t)
        ws[f"E{r}"].fill = rf; ws[f"E{r}"].alignment = aln("center"); ws[f"E{r}"].border = borda
        ws[f"G{r}"].fill = rf; ws[f"G{r}"].border = borda

    # Cabeçalho tabela
    hr = 11; ws.row_dimensions[hr].height = 22
    for ci, h in enumerate(["DATA EMPLAC.","CNPJ","CLIENTE","MODELO","CHASSI","CONCESSIONÁRIA","CIDADE"],1):
        cell = ws.cell(row=hr, column=ci); cell.value = h
        cell.font = font(True, BRANCO, 9); cell.fill = fill(PRETO)
        cell.alignment = aln(); cell.border = borda

    emp_sorted = emp_mes.sort_values("Data emplacamento").reset_index(drop=True)
    for di, drow in emp_sorted.iterrows():
        r = hr + 1 + di; ws.row_dimensions[r].height = 16
        cn = str(drow.get("CNPJ_NORM",""))
        is_dn = is_denigris(pd.Series([drow.get("Concessionário","")])).iloc[0]
        is_conc = cn in cnpjs_carteira_set and not is_dn
        row_fill = fill(AZUL_CL) if is_dn else (fill(ROSA_CL) if is_conc else fill(BRANCO if di%2==0 else CINZA_CL))
        dv = drow.get("Data emplacamento")
        ds = dv.strftime("%d/%m/%Y") if hasattr(dv,"strftime") else str(dv)[:10]
        for ci, val in enumerate([
            ds, str(drow.get("CPFCNPJPROPRIETARIO",""))[:20],
            str(drow.get("NOMEPROPRIETARIO",""))[:45], str(drow.get("Modelo",""))[:30],
            str(drow.get("Chassi",""))[:20], str(drow.get("Concessionário",""))[:35],
            str(drow.get("NO_CIDADE",""))[:20]], 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = Font(name="Arial",size=8,color="FF1A1A1A")
            cell.fill = row_fill; cell.border = borda
            cell.alignment = Alignment(horizontal="left" if ci>1 else "center", vertical="center")

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ════════════════════════════════════════════════════════════════
# SESSION STATE INIT
# ════════════════════════════════════════════════════════════════
for k, v in [("user",None),("df_area",None),("df_cart",None),
             ("df_emp_list",[]),("emp_fontes",[]),("pagina","busca"),
             ("users_db",None),("dados_carregados",False)]:
    if k not in st.session_state:
        st.session_state[k] = v

if st.session_state.users_db is None:
    st.session_state.users_db = load_users()

USERS = st.session_state.users_db

def carregar_dados_se_necessario():
    if st.session_state.user is None:
        return
    if st.session_state.get("dados_carregados"):
        return
    if st.session_state.df_cart is None:
        src = None
        for filename in ["CARTEIRA_VANS.xlsx","CARTEIRA.xlsx"]:
            src, err = load_excel_from_github(filename)
            if src:
                st.session_state["_cart_erro"] = None
                break
            else:
                st.session_state["_cart_erro"] = f"{filename}: {err}"
        if src:
            st.session_state.df_cart = load_carteira(src)
    if not st.session_state.df_emp_list:
        for arq in ["EMPLACAMENTO APP VANS.xlsx"]:
            src, err = load_excel_from_github(arq)
            if src:
                df = load_emplacamentos(src, label=arq)
                if df is not None:
                    st.session_state.df_emp_list.append(df)
            else:
                st.session_state["_emp_erros"] = [f"{arq}: {err}"]
    if st.session_state.df_cart is not None or st.session_state.df_emp_list:
        st.session_state.dados_carregados = True

carregar_dados_se_necessario()

# ════════════════════════════════════════════════════════════════
# LOGIN
# ════════════════════════════════════════════════════════════════
if st.session_state.user is None:
    st.markdown("""
    <style>
    [data-testid="stAppViewContainer"]{
        background:linear-gradient(160deg,#0a1628 0%,#0d1f3c 60%,#0a1628 100%)!important;min-height:100vh;
    }
    [data-testid="stMain"],section[data-testid="stMain"]>div{background:transparent!important;}
    .main .block-container{padding-top:0!important;padding-bottom:0!important;max-width:100%!important;}
    </style>""", unsafe_allow_html=True)

    col_l, col_c, col_r = st.columns([1, 2, 1])
    with col_c:
        st.markdown(f"""
        <div style="background:#fff;border-radius:20px;padding:32px 24px 18px 24px;
                    box-shadow:0 28px 56px rgba(0,0,0,0.32);margin-top:52px;text-align:center;">
            <div style="background:#0a1628;border-radius:14px;padding:12px 18px;
                        display:inline-block;margin-bottom:8px;">
                <span style="font-size:20px;font-weight:800;color:#fff;">Comercial De Nigris</span>
            </div>
            <div style="font-size:10px;color:#8a95b0;text-transform:uppercase;
                        letter-spacing:2px;margin-bottom:6px;">Inteligência Comercial</div>
        </div>""", unsafe_allow_html=True)

        with st.form("form_login", clear_on_submit=False):
            usuario_input = st.text_input("", placeholder="👤  Usuário", label_visibility="collapsed")
            senha_input   = st.text_input("", placeholder="🔒  Senha", type="password", label_visibility="collapsed")
            submitted     = st.form_submit_button("Entrar →", use_container_width=True)

        if submitted:
            key = usuario_input.strip().upper()
            users_fresh = load_users()
            st.session_state.users_db = users_fresh
            if not key:
                st.error("❌ Digite seu usuário")
            elif key in users_fresh:
                u = users_fresh[key]
                senha_ok = (
                    u.get("senha_hash") == hash_senha(senha_input.strip()) or
                    u.get("senha") == senha_input.strip()
                )
                if senha_ok:
                    st.session_state.user = key
                    registrar_acesso(key)
                    st.rerun()
                else:
                    st.error("❌ Senha incorreta")
            else:
                st.error(f"❌ Usuário '{key}' não encontrado")
    st.stop()

# ════════════════════════════════════════════════════════════════
# USUÁRIO LOGADO
# ════════════════════════════════════════════════════════════════
carregar_dados_se_necessario()

u_key  = st.session_state.user
u_data = USERS.get(u_key, {})
perfil = u_data.get("perfil", "vendedor")
nome   = u_data.get("nome", u_key)
cons_key = u_data.get("consultor_key", u_key)

df_area = st.session_state.df_area
df_cart = st.session_state.df_cart

_emp_list_len = len(st.session_state.df_emp_list)
if st.session_state.df_emp_list and st.session_state.get("_emp_merged_len") != _emp_list_len:
    st.session_state["_df_emp_merged"] = merge_emp(st.session_state.df_emp_list)
    st.session_state["_emp_merged_len"] = _emp_list_len
df_emp = st.session_state.get("_df_emp_merged") if st.session_state.df_emp_list else None

PAGINAS_GESTOR  = [("busca","🔍","Busca"),("emplacamentos","📍","Emplacam."),("carteira","📋","Carteira"),("painel","📊","Painel"),("gestao","📈","Gestão"),("oportunidades","🎯","Oportun."),("admin","⚙️","Admin")]
PAGINAS_GERENTE = [("busca","🔍","Busca"),("emplacamentos","📍","Emplacam."),("carteira","📋","Carteira"),("painel","📊","Painel"),("gestao","📈","Gestão"),("oportunidades","🎯","Oportun.")]
PAGINAS_VEND    = [("busca","🔍","Busca"),("emplacamentos","📍","Emplacam."),("carteira","📋","Carteira"),("oportunidades","🎯","Oportun.")]

if perfil == "gestor": PAGINAS = PAGINAS_GESTOR
elif perfil == "gerente": PAGINAS = PAGINAS_GERENTE
else: PAGINAS = PAGINAS_VEND

pagina = st.session_state.pagina
if pagina not in [p[0] for p in PAGINAS]:
    pagina = PAGINAS[0][0]
    st.session_state.pagina = pagina

# ── TOPBAR ──
sigla = nome[0].upper()
perfil_label = {"gestor":"Administrador","gerente":"Gerente","vendedor":"Consultor"}.get(perfil,"Usuário")
st.markdown(f"""
<div class="topbar">
  <div class="topbar-logo"><span>Comercial De Nigris</span></div>
  <div class="topbar-user">
    <div class="topbar-info">
      <div class="topbar-name">{nome}</div>
      <div class="topbar-role">{perfil_label}</div>
    </div>
    <div class="topbar-avatar">{sigla}</div>
  </div>
</div>""", unsafe_allow_html=True)

# ── NAVEGAÇÃO ──
nav_cols = st.columns(len(PAGINAS) + 1)
for i, (pid, picon, plabel) in enumerate(PAGINAS):
    with nav_cols[i]:
        if st.button(f"{picon} {plabel}", key=f"nav_{pid}", use_container_width=True):
            st.session_state.pagina = pid
            st.rerun()
with nav_cols[-1]:
    if st.button("🚪 Sair", key="sair"):
        st.session_state.user = None
        st.rerun()

# ════════════════════════════════════════════════════════════════
# CONTEÚDO DAS PÁGINAS
# ════════════════════════════════════════════════════════════════
st.markdown('<div class="page-wrap">', unsafe_allow_html=True)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PÁGINA: BUSCA DE CLIENTE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
if pagina == "busca":
    st.markdown("""<div class="page-header"><h1>🔍 Busca de Cliente</h1>
    <p>Pesquise por razão social, CNPJ ou placa</p></div>""", unsafe_allow_html=True)

    if df_emp is None:
        st.warning("⚠️ Dados não carregados. Verifique a página Admin.")
        st.stop()

    q = st.text_input("", placeholder="Nome, CNPJ, CPF ou Placa...", label_visibility="collapsed")
    buscar = st.button("Buscar Cliente", use_container_width=True)

    if buscar and q:
        q_strip = q.strip()
        q_norm  = norm_str(q_strip)
        q_cnpj  = re.sub(r"\D", "", q_strip)
        q_placa = q_strip.replace("-","").replace(" ","").upper()
        tokens  = [t for t in q_norm.split() if len(t) >= 2]

        nome_norm_series = norm_str_series(df_emp["NOMEPROPRIETARIO"])
        if tokens:
            mask_nome = pd.Series(True, index=df_emp.index)
            for tok in tokens:
                mask_nome = mask_nome & nome_norm_series.str.contains(tok, na=False, regex=False)
        else:
            mask_nome = pd.Series(False, index=df_emp.index)

        mask_cnpj  = df_emp["CNPJ_NORM"] == q_cnpj if len(q_cnpj) >= 11 else (df_emp["CNPJ_NORM"].str.startswith(q_cnpj) if len(q_cnpj) >= 3 else pd.Series(False, index=df_emp.index))
        mask_placa = df_emp["Placa_norm"].str.contains(q_placa, na=False, regex=False) if len(q_placa) >= 3 else pd.Series(False, index=df_emp.index)

        if mask_placa.any() and len(q_placa) >= 3 and not mask_nome.any() and not mask_cnpj.any():
            resultados = df_emp[mask_placa].sort_values("Data emplacamento", ascending=False)
        else:
            mask = mask_nome | mask_cnpj | mask_placa
            resultados = df_emp[mask].copy()
            if not resultados.empty and tokens:
                _ns = norm_str_series(resultados["NOMEPROPRIETARIO"])
                resultados["_score"] = sum(_ns.str.contains(tok, na=False, regex=False).astype(int) for tok in tokens)
                resultados = resultados.sort_values("_score", ascending=False)

        cnpjs = resultados["CNPJ_NORM"].unique()

        if len(cnpjs) == 0:
            st.warning("Nenhum cliente encontrado.")
        else:
            if len(cnpjs) > 1:
                opts = []
                for cn in cnpjs[:20]:
                    sub = df_emp[df_emp["CNPJ_NORM"] == cn]
                    opts.append(f"{sub['NOMEPROPRIETARIO'].iloc[0]} — {sub['NO_CIDADE'].iloc[0]} ({len(sub)} empl.)")
                sel = st.selectbox("Selecione o cliente:", opts)
                cnpj_sel = cnpjs[opts.index(sel)]
            else:
                cnpj_sel = cnpjs[0]

            edf  = df_emp[df_emp["CNPJ_NORM"] == cnpj_sel].copy()
            esrt = edf.sort_values("Data emplacamento", ascending=False)
            last = esrt.iloc[0].to_dict()
            total_emp = len(edf)
            datas = edf["Data emplacamento"].dropna().tolist()
            pred_label, pred_date = calc_prediction(datas)
            nigris_cnt = int(is_denigris(edf["Concessionário"]).sum())

            cart_row = None
            if df_cart is not None:
                cr = df_cart[df_cart["CNPJ_NORM"] == cnpj_sel]
                cart_row = cr.iloc[0].to_dict() if not cr.empty else None

            vendedores_ativos = get_vendedores_ativos()
            consultor_resp = get_vendedor_final(cnpj_sel, cart_row, vendedores_ativos)

            nome_exib = safe_str(last.get("NOME_FANTASIA",""))
            if nome_exib == "—": nome_exib = safe_str(last.get("NOMEPROPRIETARIO",""))
            cls_raw = safe_str(cart_row.get("Classificação Mercedes","")) if cart_row else "—"
            badge_h = f'<span class="badge {badge_class(cls_raw)}">{cls_raw}</span>' if cls_raw != "—" else ""

            resp_h = ""
            if consultor_resp and consultor_resp != "—":
                is_da_carteira = cart_row and safe_str(cart_row.get("VENDEDOR","")) == consultor_resp
                icon = "👤" if is_da_carteira else "⚖️"
                bg = "#e8f0ff" if is_da_carteira else "#f0f0f0"
                tc = "#0044aa" if is_da_carteira else "#555555"
                resp_h = f'<span style="background:{bg};color:{tc};padding:3px 10px;border-radius:20px;font-size:10px;font-weight:700;">{icon} {consultor_resp}</span>'

            pred_mes = pred_label.split(' de ')[0][:3] if pred_label else "—"
            pred_ano = ("Prev. "+pred_label.split(' de ')[1]) if pred_label else "Sem previsão"

            st.markdown(f"""
            <div class="client-hero">
                <div class="hero-name">{nome_exib}</div>
                <div class="hero-cnpj">{safe_str(last.get("CPFCNPJPROPRIETARIO",""))}</div>
                <div style="margin-top:7px;">{badge_h}</div>
                <div style="margin-top:7px;">{resp_h}</div>
                <div class="hero-stats">
                    <div class="hero-stat">
                        <div class="hero-stat-val">{total_emp}</div>
                        <div class="hero-stat-lbl">Emplacamentos</div>
                    </div>
                    <div class="hero-stat">
                        <div class="hero-stat-val" style="color:#44cc88;">{nigris_cnt}</div>
                        <div class="hero-stat-lbl">De Nigris</div>
                    </div>
                    <div class="hero-stat">
                        <div class="hero-stat-val">{pred_mes}</div>
                        <div class="hero-stat-lbl">{pred_ano}</div>
                    </div>
                </div>
            </div>""", unsafe_allow_html=True)

            tab_cad, tab_contatos, tab_socios, tab_hist, tab_receita = st.tabs(["📋 Cadastro","📞 Contatos","🤝 Sócios","📈 Histórico","🏢 Receita Federal"])

            with tab_cad:
                end_parts = [safe_str(last.get(c,""),"") for c in ["TP_LOGR","NO_LOGR","NU_LOGR","NO_COMPL","NO_BAIRRO"]]
                endereco  = " ".join(p for p in end_parts if p and p!="—").strip() or "—"
                cidade_uf = f"{safe_str(last.get('NO_CIDADE',''))} - {safe_str(last.get('SG_ESTADO',''))}"
                cep_exib  = norm_cep(last.get("NU_CEP",""))
                if len(cep_exib) == 8: cep_exib = f"{cep_exib[:5]}-{cep_exib[5:]}"

                ultima_data_str = "—"; ultimo_modelo = "—"; foi_nigris = "—"
                if not esrt.empty:
                    ud = pd.to_datetime(esrt["Data emplacamento"].iloc[0], errors="coerce")
                    ultima_data_str = ud.strftime("%d/%m/%Y") if pd.notna(ud) else "—"
                    ultimo_modelo = safe_str(esrt["Modelo"].iloc[0])
                    ultima_conc   = safe_str(esrt["Concessionário"].iloc[0])
                    foi_nigris = "✅ Comercial De Nigris" if is_denigris(pd.Series([ultima_conc])).iloc[0] else f"⚠️ {ultima_conc[:40]}"

                infos = [
                    ("Razão Social", safe_str(last.get("NOMEPROPRIETARIO",""))),
                    ("Nome Fantasia", safe_str(last.get("NOME_FANTASIA",""))),
                    ("CNPJ / CPF", safe_str(last.get("CPFCNPJPROPRIETARIO",""))),
                    ("📅 Últ. Compra", ultima_data_str),
                    ("🚚 Últ. Modelo", ultimo_modelo),
                    ("🏢 Últ. Conc.", foi_nigris),
                    ("Endereço", endereco),
                    ("Cidade / UF", cidade_uf),
                    ("CEP", cep_exib or "—"),
                    ("Atividade", safe_atividade(last)),
                    ("Nat. Jurídica", safe_str(last.get("NATUREZA_JURIDICA",""))),
                    ("Situação", safe_str(last.get("SITUACAO_RECEITA",""))),
                ]
                if cart_row:
                    infos += [
                        ("Segmento", safe_str(cart_row.get("Segmento principal do cliente",""))),
                        ("Classificação", cls_raw),
                        ("Vendedor", safe_str(cart_row.get("VENDEDOR",""))),
                    ]

                import urllib.parse as _urlp
                end_enc = _urlp.quote(f"{endereco} {cidade_uf}".strip())
                gmaps_url = f"https://www.google.com/maps/search/?api=1&query={end_enc}"
                waze_url  = f"https://waze.com/ul?q={end_enc}"

                st.markdown('<div class="info-card">', unsafe_allow_html=True)
                for lbl, val in infos:
                    if val and val != "—":
                        if lbl in ("Endereço","Cidade / UF"):
                            nav_html = (f'<div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;">'
                                f'<span>{val}</span>'
                                f'<a href="{gmaps_url}" target="_blank" style="background:#e8f0ff;color:#1a73e8;padding:2px 7px;border-radius:16px;font-size:9px;font-weight:700;text-decoration:none;white-space:nowrap;">🗺️ Maps</a>'
                                f'<a href="{waze_url}" target="_blank" style="background:#f0eaff;color:#7c3aed;padding:2px 7px;border-radius:16px;font-size:9px;font-weight:700;text-decoration:none;white-space:nowrap;">🚗 Waze</a>'
                                f'</div>')
                            st.markdown(f'<div class="info-row"><div class="info-label">{lbl}</div><div class="info-value">{nav_html}</div></div>', unsafe_allow_html=True)
                        else:
                            st.markdown(f'<div class="info-row"><div class="info-label">{lbl}</div><div class="info-value">{val}</div></div>', unsafe_allow_html=True)
                st.markdown('</div>', unsafe_allow_html=True)

            with tab_contatos:
                st.markdown('<div class="sec-title">📞 Contatos Disponíveis</div>', unsafe_allow_html=True)
                phones = []
                msg_wa = "Olá! Entro em contato da Comercial De Nigris para apresentar nossas soluções."

                for i in range(1,6):
                    t = format_tel(last.get(f"DDD{i}"), last.get(f"TELEFONE{i}"))
                    n = make_fone_num(last.get(f"DDD{i}"), last.get(f"TELEFONE{i}"))
                    if t: phones.append(("fixo", t, n, "Emplacamento"))
                for i in range(1,4):
                    t = format_tel(last.get(f"DDD_CELULAR{i}"), last.get(f"CELULAR{i}"))
                    n = make_fone_num(last.get(f"DDD_CELULAR{i}"), last.get(f"CELULAR{i}"))
                    if t: phones.append(("cel", t, n, "Emplacamento"))
                email_emp = safe_str(last.get("EMAIL",""))
                email_cart = ""
                if cart_row:
                    for campo in ["Telefone Residencial","Telefone Comercial"]:
                        v = safe_str(cart_row.get(campo,""))
                        if v != "—":
                            phones.append(("fixo", v, "55"+re.sub(r"\D","",v), "Carteira"))
                    cel_c = safe_str(cart_row.get("Celular",""))
                    if cel_c != "—":
                        phones.append(("cel", cel_c, "55"+re.sub(r"\D","",cel_c), "Carteira"))
                    email_cart = safe_str(cart_row.get("E-mail",""))
                    if email_cart != "—":
                        st.markdown(f'<a href="mailto:{email_cart}" class="contact-btn btn-email" style="display:flex;flex-direction:row;align-items:center;gap:10px;padding:11px 14px;border-radius:12px;margin-bottom:9px;"><span style="font-size:20px;">✉️</span><div><div style="font-weight:700;font-size:12px;">{email_cart}</div><div style="font-size:10px;color:#8a95b0;">E-mail · Carteira</div></div></a>', unsafe_allow_html=True)
                if email_emp != "—" and email_emp != email_cart:
                    st.markdown(f'<a href="mailto:{email_emp}" class="contact-btn btn-email" style="display:flex;flex-direction:row;align-items:center;gap:10px;padding:11px 14px;border-radius:12px;margin-bottom:9px;"><span style="font-size:20px;">✉️</span><div><div style="font-weight:700;font-size:12px;">{email_emp}</div><div style="font-size:10px;color:#8a95b0;">E-mail · Base</div></div></a>', unsafe_allow_html=True)

                seen = set()
                for tipo, fmt, num, fonte in phones:
                    if fmt in seen: continue
                    seen.add(fmt)
                    wa_url  = f"https://wa.me/{num}?text={msg_wa}"
                    tel_url = f"tel:+{num}"
                    if tipo == "cel":
                        st.markdown(f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:7px;margin-bottom:7px;"><a href="{wa_url}" target="_blank" class="contact-btn btn-whatsapp" style="display:flex;flex-direction:row;align-items:center;gap:7px;padding:9px 12px;border-radius:10px;"><span style="font-size:18px;">💬</span><div><div style="font-weight:700;font-size:11px;">{fmt}</div><div style="font-size:9px;">WhatsApp · {fonte}</div></div></a><a href="{tel_url}" class="contact-btn btn-phone" style="display:flex;flex-direction:row;align-items:center;gap:7px;padding:9px 12px;border-radius:10px;"><span style="font-size:18px;">📞</span><div><div style="font-weight:700;font-size:11px;">{fmt}</div><div style="font-size:9px;">Ligar · {fonte}</div></div></a></div>', unsafe_allow_html=True)
                    else:
                        st.markdown(f'<a href="{tel_url}" class="contact-btn btn-phone" style="display:flex;flex-direction:row;align-items:center;gap:7px;padding:9px 12px;border-radius:10px;margin-bottom:7px;"><span style="font-size:18px;">📞</span><div><div style="font-weight:700;font-size:11px;">{fmt}</div><div style="font-size:9px;">Fixo · {fonte}</div></div></a>', unsafe_allow_html=True)

                if not phones and email_emp == "—" and email_cart == "—":
                    st.info("Nenhum contato disponível.")

                st.markdown('<div class="sec-title">🏷️ Preferências de Compra</div>', unsafe_allow_html=True)
                marcas  = get_modes(edf["Marca"], top=3)
                modelos = get_modes(edf["Modelo"], top=3)
                concs   = get_modes(edf["Concessionário"], top=2)
                st.markdown('<div class="info-card">', unsafe_allow_html=True)
                for m in marcas:
                    cnt = len(edf[edf["Marca"]==m])
                    st.markdown(f'<div class="info-row"><div class="info-label">Marca</div><div class="info-value"><strong>{m}</strong> <span style="color:#8a95b0;">({cnt}x)</span></div></div>', unsafe_allow_html=True)
                for mod in modelos:
                    cnt = len(edf[edf["Modelo"]==mod])
                    st.markdown(f'<div class="info-row"><div class="info-label">Modelo</div><div class="info-value">{mod} ({cnt}x)</div></div>', unsafe_allow_html=True)
                for conc in concs:
                    icon = "✅" if any(n in conc.upper() for n in NOMES_DENIGRIS) else "⚠️"
                    st.markdown(f'<div class="info-row"><div class="info-label">Concessionária</div><div class="info-value">{icon} {conc[:50]}</div></div>', unsafe_allow_html=True)
                st.markdown('</div>', unsafe_allow_html=True)

            with tab_socios:
                socios = []
                for i in range(1,4):
                    ns = safe_str(last.get(f"NOME_SOCIO_DIRETOR{i}",""), "")
                    if ns and ns != "—":
                        cel_s  = format_tel(last.get(f"DDD1_CEL_SOCIO{i}"), last.get(f"TEL1_CEL_SOCIO{i}"))
                        cel_n  = make_fone_num(last.get(f"DDD1_CEL_SOCIO{i}"), last.get(f"TEL1_CEL_SOCIO{i}"))
                        email_s= safe_str(last.get(f"EMAIL_SOCIO{i}",""))
                        cargo_s= safe_str(last.get(f"CARGO{i}",""))
                        socios.append({"nome":ns,"cargo":cargo_s,"email":email_s,"cel":cel_s,"cel_n":cel_n})
                if socios:
                    for s in socios:
                        btns = ""
                        if s["cel"] and s["cel"] != "—":
                            msg_s = "Olá! Entro em contato da Comercial De Nigris."
                            btns += f'<a href="https://wa.me/{s["cel_n"]}?text={msg_s}" target="_blank" style="display:inline-flex;align-items:center;gap:4px;background:#e8fff2;color:#007a3d;padding:4px 10px;border-radius:10px;font-size:11px;font-weight:700;text-decoration:none;margin-right:5px;">💬 {s["cel"]}</a>'
                            btns += f'<a href="tel:+{s["cel_n"]}" style="display:inline-flex;align-items:center;gap:4px;background:#e8f4ff;color:#0055bb;padding:4px 10px;border-radius:10px;font-size:11px;font-weight:700;text-decoration:none;margin-right:5px;">📞 Ligar</a>'
                        if s["email"] != "—":
                            btns += f'<a href="mailto:{s["email"]}" style="display:inline-flex;align-items:center;gap:4px;background:#fff8e8;color:#a06000;padding:4px 10px;border-radius:10px;font-size:11px;font-weight:700;text-decoration:none;">✉️ E-mail</a>'
                        st.markdown(f'<div class="socio-card"><div class="socio-name">{s["nome"]}</div><div class="socio-role">{s["cargo"]}</div><div style="font-size:11px;color:#6b7a99;margin-top:5px;">{s["email"]}</div><div style="margin-top:9px;">{btns}</div></div>', unsafe_allow_html=True)
                else:
                    st.info("Dados societários não disponíveis na base.")

            with tab_hist:
                h1, h2, h3, h4 = st.columns(4)
                total_h = len(edf); nigris_h = int(is_denigris(edf["Concessionário"]).sum())
                h1.metric("Total", total_h); h2.metric("De Nigris", nigris_h)
                h3.metric("Concorrência", total_h-nigris_h); h4.metric("Anos ativo", edf["Ano"].nunique())

                hist = edf.groupby("Ano").size().reset_index(name="Qtd")
                hist["Ano"] = hist["Ano"].astype(str)
                cores = ["#c8a84b" if str(a)==str(hist["Ano"].max()) else "#0a1628" for a in hist["Ano"]]
                fig = go.Figure(go.Bar(x=hist["Ano"], y=hist["Qtd"], marker_color=cores,
                    marker_line_color="#c8a84b", marker_line_width=1.2,
                    text=hist["Qtd"], textposition="outside"))
                fig.update_layout(plot_bgcolor="#fff", paper_bgcolor="#fff", font_color="#4a5568",
                    height=200, margin=dict(t=20,b=10,l=10,r=10),
                    xaxis=dict(showgrid=False), yaxis=dict(showgrid=True, gridcolor="#f0f2f7"))
                st.plotly_chart(fig, use_container_width=True)

                cols_hist = [c for c in ["Data emplacamento","Placa","Modelo","Marca","Concessionário","NO_CIDADE"] if c in esrt.columns]
                det = esrt[cols_hist].copy()
                det["Data emplacamento"] = pd.to_datetime(det["Data emplacamento"], errors="coerce").dt.strftime("%d/%m/%Y")
                det["De Nigris"] = is_denigris(esrt["Concessionário"]).map({True:"✅", False:"—"})
                det = det.rename(columns={"Data emplacamento":"Data","Concessionário":"Concessionária","NO_CIDADE":"Cidade"})
                st.dataframe(det, use_container_width=True, hide_index=True)

            with tab_receita:
                cnpj_rf = re.sub(r"\D", "", str(cnpj_sel))
                cnpj_fmt_rf = f"{cnpj_rf[:2]}.{cnpj_rf[2:5]}.{cnpj_rf[5:8]}/{cnpj_rf[8:12]}-{cnpj_rf[12:14]}" if len(cnpj_rf)==14 else cnpj_rf
                nome_rf = safe_str(last.get("NOMEPROPRIETARIO","—"))

                col_btn_rf, col_info_rf = st.columns([2,3])
                with col_btn_rf:
                    btn_rf = st.button("🔎 Consultar Receita Federal", use_container_width=True, key="btn_receita")
                with col_info_rf:
                    st.markdown(f'<div style="padding:8px 10px;background:#f0f8ff;border-radius:8px;font-size:11px;color:#4a5568;word-break:break-all;">CNPJ: <strong>{cnpj_fmt_rf}</strong><br>{nome_rf[:50]}</div>', unsafe_allow_html=True)

                if btn_rf:
                    if len(cnpj_rf) != 14:
                        st.warning("⚠️ CNPJ inválido para consulta (precisa de 14 dígitos).")
                    else:
                        with st.spinner("Consultando base pública da Receita Federal..."):
                            dados_rf, erro_rf = consultar_receita_federal(cnpj_rf)
                        if dados_rf:
                            st.session_state[f"rf_{cnpj_rf}"] = dados_rf
                            st.success("✅ Dados obtidos com sucesso!")
                        else:
                            # Mostrar fallback com link externo
                            link_serpro = f"https://servicos.receita.fazenda.gov.br/servicos/cnpjreva/cnpjreva_solicitacao.asp"
                            link_cnpjws = f"https://cnpj.ws/cnpj/{cnpj_rf}"
                            link_brasil  = f"https://brasilapi.com.br/api/cnpj/v1/{cnpj_rf}"
                            st.markdown(f"""
                            <div class="alert-yellow">
                            ⚠️ <strong>Consulta automática temporariamente indisponível</strong><br>
                            <small style="color:#666;">Erro: {erro_rf[:120] if erro_rf else 'timeout'}</small><br><br>
                            Consulte manualmente em um dos links abaixo:<br>
                            <div style="display:flex;gap:8px;flex-wrap:wrap;margin-top:8px;">
                              <a href="{link_cnpjws}" target="_blank" style="background:#0a1628;color:#c8a84b;padding:5px 12px;border-radius:8px;font-size:11px;font-weight:700;text-decoration:none;">🔗 CNPJ.ws</a>
                              <a href="{link_serpro}" target="_blank" style="background:#1a6b3a;color:#fff;padding:5px 12px;border-radius:8px;font-size:11px;font-weight:700;text-decoration:none;">🏛️ Receita Federal</a>
                              <a href="https://solucoes.receita.fazenda.gov.br/servicos/cnpjreva/Cnpjreva_Solicitacao.asp?cnpj={cnpj_rf}" target="_blank" style="background:#0044aa;color:#fff;padding:5px 12px;border-radius:8px;font-size:11px;font-weight:700;text-decoration:none;">📋 Comprovante RF</a>
                            </div>
                            </div>""", unsafe_allow_html=True)

                dados_rf = st.session_state.get(f"rf_{cnpj_rf}")
                if dados_rf:
                    try:
                        p = parse_receita_federal(dados_rf)
                        cor_sit = "#1E7E34" if "ATIVA" in str(p["situacao"]).upper() else "#C0392B"

                        st.markdown(f"""
                        <div class="rf-card">
                            <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:14px;flex-wrap:wrap;gap:8px;">
                                <div>
                                    <div style="font-size:14px;font-weight:700;color:#0a1628;">{p["razao"]}</div>
                                    {"<div style='font-size:11px;color:#8a95b0;margin-top:2px;'>"+p["fantasia"]+"</div>" if p["fantasia"] not in ["—",""] and p["fantasia"]!=p["razao"] else ""}
                                    <div style="font-size:10px;color:#8a95b0;margin-top:3px;font-family:monospace;">{cnpj_fmt_rf}</div>
                                </div>
                                <div style="background:{cor_sit};color:#fff;padding:4px 12px;border-radius:16px;font-size:11px;font-weight:600;white-space:nowrap;">
                                    ● {p["situacao"]}
                                </div>
                            </div>
                            <div class="rf-grid">
                                <div class="rf-item"><div class="rf-item-label">Data de Abertura</div><div class="rf-item-value">{p["abertura"]}</div></div>
                                <div class="rf-item"><div class="rf-item-label">Capital Social</div><div class="rf-item-value">{p["capital"]}</div></div>
                                <div class="rf-item"><div class="rf-item-label">Porte</div><div class="rf-item-value">{p["porte"]}</div></div>
                                <div class="rf-item"><div class="rf-item-label">Natureza Jurídica</div><div class="rf-item-value">{p["nat_jur"]}</div></div>
                            </div>
                            <div class="rf-item" style="margin-bottom:10px;">
                                <div class="rf-item-label">Atividade Principal</div>
                                <div class="rf-item-value">{str(p["atividade"])[:100]}</div>
                            </div>
                            <div class="rf-item">
                                <div class="rf-item-label">Endereço</div>
                                <div class="rf-item-value">{p["end_rf"]}{"<br><span style='color:#8a95b0;font-size:10px;'>"+p["cidade_rf"]+" — "+p["uf_rf"]+" · CEP "+p["cep_rf"]+"</span>" if p["cidade_rf"] not in ["—",""] else ""}</div>
                            </div>
                        </div>""", unsafe_allow_html=True)

                        st.markdown('<div class="sec-title">🤝 Quadro Societário</div>', unsafe_allow_html=True)
                        socios_rf = p["socios"]
                        if socios_rf:
                            for s in socios_rf:
                                nome_s = s.get("nome") or s.get("nome_socio","—")
                                qual_s = s.get("qualificacao_socio",{}).get("descricao","") if isinstance(s.get("qualificacao_socio"),dict) else s.get("qual","Sócio")
                                cpf_s  = s.get("cpf_representante_legal") or s.get("cnpj_cpf_do_socio","")
                                desde_s= s.get("data_entrada_sociedade","—")
                                st.markdown(f'<div style="background:#f8fafc;border:1px solid #e2e8f0;border-left:3px solid #c8a84b;border-radius:8px;padding:10px 14px;margin-bottom:7px;"><div style="font-weight:700;font-size:12px;color:#0a1628;">👤 {nome_s}</div><div style="font-size:10px;color:#8a95b0;margin-top:2px;">{qual_s} · Desde {desde_s}</div><div style="font-size:10px;color:#4a5568;margin-top:2px;font-family:monospace;">{cpf_s or "CPF não divulgado"}</div></div>', unsafe_allow_html=True)
                        else:
                            st.info("Nenhum sócio encontrado na base pública.")
                        st.markdown('<div style="font-size:9px;color:#b0b8cc;margin-top:12px;text-align:center;">Fonte: API pública CNPJ.ws / BrasilAPI · Receita Federal do Brasil</div>', unsafe_allow_html=True)
                    except Exception as e_rf:
                        st.error(f"Erro ao processar dados: {e_rf}")
                else:
                    st.markdown("""
                    <div style="text-align:center;padding:40px 16px;color:#8a95b0;">
                        <div style="font-size:44px;margin-bottom:14px;">🏢</div>
                        <div style="font-size:14px;font-weight:600;color:#4a5568;margin-bottom:7px;">Dados da Receita Federal</div>
                        <div style="font-size:12px;">Clique em <strong>Consultar Receita Federal</strong><br>para buscar situação cadastral e sócios.</div>
                    </div>""", unsafe_allow_html=True)

    elif buscar and not q:
        st.warning("Digite algo para buscar.")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PÁGINA: EMPLACAMENTOS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
elif pagina == "emplacamentos":
    st.markdown("""<div class="page-header"><h1>📍 Meus Emplacamentos</h1>
    <p>Visão completa da sua área e carteira</p></div>""", unsafe_allow_html=True)

    if df_emp is None:
        st.warning("⚠️ Dados não carregados."); st.stop()

    today = pd.Timestamp.now()

    if perfil in ("gestor","gerente"):
        vendedores_disp = get_vendedores_ativos()
        sel_cons = st.selectbox("Consultor:", ["Todos"] + vendedores_disp)
    else:
        sel_cons = nome  # vendedor vê pelo nome do usuário

    # Filtro período
    datas_validas = df_emp.dropna(subset=["Data emplacamento"])
    datas_validas = datas_validas[datas_validas["Data emplacamento"] <= today]
    ultimo_reg = datas_validas["Data emplacamento"].max() if not datas_validas.empty else today
    ano_default = int(ultimo_reg.year); mes_default = int(ultimo_reg.month)
    anos_disp = sorted([int(a) for a in df_emp["Ano"].dropna().unique() if int(a) <= today.year], reverse=True)
    if not anos_disp: anos_disp = sorted([int(a) for a in df_emp["Ano"].dropna().unique()], reverse=True)

    fc1, fc2 = st.columns(2)
    with fc1:
        idx_ano = anos_disp.index(ano_default) if ano_default in anos_disp else 0
        sel_ano = st.selectbox("Ano:", anos_disp, index=idx_ano)
    with fc2:
        meses_do_ano = sorted(df_emp[df_emp["Ano"]==sel_ano]["Mes"].dropna().astype(int).unique().tolist())
        if not meses_do_ano: meses_do_ano = list(range(1,13))
        mes_labels = [MESES_PT[m] for m in meses_do_ano]
        mes_def_lbl = MESES_PT[mes_default] if sel_ano == ano_default and mes_default in meses_do_ano else mes_labels[-1]
        idx_mes = mes_labels.index(mes_def_lbl) if mes_def_lbl in mes_labels else len(mes_labels)-1
        sel_mes_lbl = st.selectbox("Mês:", mes_labels, index=idx_mes)
        sel_mes = meses_do_ano[mes_labels.index(sel_mes_lbl)]

    # CNPJs da carteira
    if df_cart is not None:
        todos_cnpjs_cart = set(df_cart["CNPJ_NORM"].dropna().unique())
        if sel_cons != "Todos":
            sel_cons_norm = norm_str(sel_cons)
            cnpjs_carteira_set = set(df_cart[
                df_cart["VENDEDOR"].apply(norm_str) == sel_cons_norm
            ]["CNPJ_NORM"].dropna().unique())
        else:
            cnpjs_carteira_set = todos_cnpjs_cart
    else:
        todos_cnpjs_cart = set()
        cnpjs_carteira_set = set()

    # Todos os emplacamentos do período
    emp_periodo = df_emp[(df_emp["Ano"]==sel_ano) & (df_emp["Mes"]==sel_mes)].copy()

    # Atribuição dinâmica: cada CNPJ → um vendedor fixo e determinístico
    vendedores_ativos = get_vendedores_ativos()

    @st.cache_data(ttl=300, show_spinner=False)
    def calcular_atribuicoes(cnpjs_tuple, vendedores_tuple, cart_json):
        """Cache de atribuições para evitar recalcular a cada interação."""
        cart_dict = json.loads(cart_json) if cart_json else {}
        vends = list(vendedores_tuple)
        resultado = {}
        for cnpj in cnpjs_tuple:
            v_cart = cart_dict.get(cnpj, "")
            if v_cart and v_cart not in ["—","nan","None",""]:
                resultado[cnpj] = v_cart
            else:
                if vends:
                    h = int(hashlib.md5(str(cnpj).encode()).hexdigest(), 16)
                    resultado[cnpj] = vends[h % len(vends)]
                else:
                    resultado[cnpj] = None
        return resultado

    cnpjs_no_periodo = tuple(sorted(emp_periodo["CNPJ_NORM"].unique()))
    # Preparar dict cnpj→vendedor da carteira
    if df_cart is not None:
        cart_vend_dict = df_cart.drop_duplicates(subset=["CNPJ_NORM"]).set_index("CNPJ_NORM")["VENDEDOR"].to_dict()
    else:
        cart_vend_dict = {}
    cart_json = json.dumps(cart_vend_dict)

    mapa_vend = calcular_atribuicoes(cnpjs_no_periodo, tuple(vendedores_ativos), cart_json)
    emp_periodo["VENDEDOR_ATRIBUIDO"] = emp_periodo["CNPJ_NORM"].map(mapa_vend)

    # Filtrar por vendedor selecionado
    if sel_cons != "Todos":
        emp_mes = emp_periodo[emp_periodo["VENDEDOR_ATRIBUIDO"] == sel_cons].copy()
    else:
        emp_mes = emp_periodo.copy()

    if emp_mes.empty:
        total_geral = len(emp_periodo)
        if total_geral == 0:
            st.info(f"Não há emplacamentos em {sel_mes_lbl}/{sel_ano} na base de dados.")
        else:
            st.info(f"Nenhum emplacamento atribuído a este consultor em {sel_mes_lbl}/{sel_ano}.")

    # Q1: Carteira que foi para concorrência
    q1_df = emp_mes[emp_mes["CNPJ_NORM"].isin(cnpjs_carteira_set) & ~is_denigris(emp_mes["Concessionário"])].copy() if cnpjs_carteira_set else pd.DataFrame()

    # Q2: Área sem cadastro (não está em NENHUMA carteira)
    q2_df = emp_mes[~emp_mes["CNPJ_NORM"].isin(todos_cnpjs_cart)].copy()

    # Q3: De Nigris
    q3_df = emp_mes[is_denigris(emp_mes["Concessionário"])].copy()

    # Q4: Top 3 do período
    top_hist = emp_mes.groupby(["CNPJ_NORM","NOMEPROPRIETARIO","NO_CIDADE"]).agg(
        Total=("Chassi","count"),
        Nigris=("Concessionário", lambda x: is_denigris(x).sum()),
    ).reset_index().sort_values("Total", ascending=False).head(3).reset_index(drop=True)

    st.markdown(f'<div class="sec-title">📊 {sel_mes_lbl} / {sel_ano}</div>', unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown(f'<div class="quadrant"><div class="quadrant-header"><div><div class="quadrant-title">🔴 Carteira → Concorrência</div><div style="font-size:10px;color:#8a95b0;margin-top:1px;">Sua carteira que comprou de outro</div></div><div class="quadrant-count q-red">{len(q1_df["CNPJ_NORM"].unique()) if not q1_df.empty else 0}</div></div></div>', unsafe_allow_html=True)
        if not q1_df.empty:
            _cols = [c for c in ["NOMEPROPRIETARIO","Modelo","Data emplacamento","Concessionário"] if c in q1_df.columns]
            q1s = q1_df.sort_values("Data emplacamento", ascending=False)[_cols].copy()
            q1s = q1s.rename(columns={"NOMEPROPRIETARIO":"Cliente","Data emplacamento":"Data","Concessionário":"Concessionária"})
            q1s["Data"] = pd.to_datetime(q1s["Data"], errors="coerce").dt.strftime("%d/%m/%Y")
            st.dataframe(q1s.head(10), use_container_width=True, hide_index=True)

    with c2:
        q2_clientes = len(q2_df["CNPJ_NORM"].unique()) if not q2_df.empty else 0
        st.markdown(f'<div class="quadrant"><div class="quadrant-header"><div><div class="quadrant-title">🟡 Área sem Cadastro</div><div style="font-size:10px;color:#8a95b0;margin-top:1px;">Emplacaram na sua área, sem carteira</div></div><div class="quadrant-count q-yellow">{q2_clientes}</div></div></div>', unsafe_allow_html=True)
        if not q2_df.empty:
            _cols = [c for c in ["NOMEPROPRIETARIO","Modelo","Data emplacamento","NO_CIDADE","Concessionário"] if c in q2_df.columns]
            q2s = q2_df.sort_values("Data emplacamento", ascending=False)[_cols].copy()
            q2s = q2s.rename(columns={"NOMEPROPRIETARIO":"Cliente","Data emplacamento":"Data","NO_CIDADE":"Cidade","Concessionário":"Concessionária"})
            q2s["Data"] = pd.to_datetime(q2s["Data"], errors="coerce").dt.strftime("%d/%m/%Y")
            st.dataframe(q2s.head(10), use_container_width=True, hide_index=True)

    c3, c4 = st.columns(2)
    with c3:
        st.markdown(f'<div class="quadrant"><div class="quadrant-header"><div><div class="quadrant-title">🟢 Compraram na De Nigris</div><div style="font-size:10px;color:#8a95b0;margin-top:1px;">Vendas realizadas por nós</div></div><div class="quadrant-count q-green">{len(q3_df["CNPJ_NORM"].unique()) if not q3_df.empty else 0}</div></div></div>', unsafe_allow_html=True)
        if not q3_df.empty:
            _cols = [c for c in ["NOMEPROPRIETARIO","Modelo","Data emplacamento","NO_CIDADE"] if c in q3_df.columns]
            q3s = q3_df.sort_values("Data emplacamento", ascending=False)[_cols].copy()
            q3s = q3s.rename(columns={"NOMEPROPRIETARIO":"Cliente","Data emplacamento":"Data","NO_CIDADE":"Cidade"})
            q3s["Data"] = pd.to_datetime(q3s["Data"], errors="coerce").dt.strftime("%d/%m/%Y")
            st.dataframe(q3s.head(10), use_container_width=True, hide_index=True)

    with c4:
        st.markdown(f'<div class="quadrant"><div class="quadrant-header"><div><div class="quadrant-title">🏆 Top 3 do Período</div><div style="font-size:10px;color:#8a95b0;margin-top:1px;">Maiores compradores no mês</div></div><div class="quadrant-count q-blue">{len(top_hist)}</div></div></div>', unsafe_allow_html=True)
        medals = ["🥇","🥈","🥉"]
        for i, row in top_hist.iterrows():
            pct_n = round(row["Nigris"]/row["Total"]*100) if row["Total"] > 0 else 0
            cor = "#007030" if pct_n >= 50 else "#a02020"
            nm = str(row['NOMEPROPRIETARIO'])
            st.markdown(f'<div style="background:#f8f9fc;border-radius:10px;padding:9px 11px;margin-bottom:7px;display:flex;align-items:center;gap:9px;"><span style="font-size:18px;">{medals[i]}</span><div style="flex:1;"><div style="font-size:11px;font-weight:800;color:#0a1628;">{nm[:28]}{"..." if len(nm)>28 else ""}</div><div style="font-size:9px;color:#8a95b0;">📍 {safe_str(row["NO_CIDADE"])} · <span style="color:{cor};font-weight:700;">{int(row["Total"])} veículos</span></div></div></div>', unsafe_allow_html=True)

    # Botão Relatório
    st.markdown("<br>", unsafe_allow_html=True)
    col_btn, col_info = st.columns([2,3])
    with col_btn:
        gerar = st.button(f"📊 Gerar Relatório — {sel_mes_lbl}/{sel_ano}", use_container_width=True, type="primary")
    with col_info:
        st.markdown(f'<div style="padding:8px 10px;background:#f0f8ff;border-radius:9px;font-size:11px;color:#4a5568;">Relatório XLSX com resumo, top 3 e lista colorida por grupo.</div>', unsafe_allow_html=True)

    if gerar:
        if emp_mes.empty:
            st.warning("Sem emplacamentos no período.")
        else:
            with st.spinner("Gerando..."):
                nome_cons = sel_cons if sel_cons != "Todos" else "Todos"
                buf_rel = gerar_relatorio_emplacamento(
                    emp_mes, emp_mes, cnpjs_carteira_set, todos_cnpjs_cart,
                    sel_mes_lbl, sel_ano, nome_cons
                )
            st.download_button(
                f"📥 Baixar Relatório_{sel_mes_lbl}_{sel_ano}.xlsx",
                data=buf_rel,
                file_name=f"Relatorio_Emplacamento_{sel_mes_lbl}_{sel_ano}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            st.success(f"✅ {len(emp_mes)} emplacamentos no relatório!")

    mostrar_todos = st.toggle("Ver todos os emplacamentos do mês", value=False)
    if mostrar_todos:
        if not emp_mes.empty:
            _cols = [c for c in ["NOMEPROPRIETARIO","Placa","Modelo","Marca","Concessionário","NO_CIDADE","Data emplacamento"] if c in emp_mes.columns]
            det = emp_mes[_cols].copy()
            det = det.rename(columns={"NOMEPROPRIETARIO":"Cliente","Concessionário":"Concessionária","NO_CIDADE":"Cidade","Data emplacamento":"Data"})
            det["Data"] = pd.to_datetime(det["Data"], errors="coerce").dt.strftime("%d/%m/%Y")
            st.dataframe(det, use_container_width=True, hide_index=True)
            buf = BytesIO(); det.to_excel(buf, index=False, engine="openpyxl"); buf.seek(0)
            st.download_button("📥 Exportar XLSX", buf, file_name=f"emplacamentos_{sel_mes_lbl}_{sel_ano}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.info("Sem emplacamentos no período.")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PÁGINA: CARTEIRA
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
elif pagina == "carteira":
    st.markdown("""<div class="page-header"><h1>📋 Carteira</h1>
    <p>Análise da sua carteira de clientes</p></div>""", unsafe_allow_html=True)

    if df_cart is None:
        st.warning("⚠️ Arquivo de Carteira não carregado."); st.stop()

    today = pd.Timestamp.now()
    vendedores_ativos = get_vendedores_ativos()

    # Carteira com atribuição dinâmica incluída
    df_dist = df_cart.copy()
    if df_emp is not None:
        nomes_emp = df_emp.sort_values("Data emplacamento").groupby("CNPJ_NORM").last()[["NOMEPROPRIETARIO","CPFCNPJPROPRIETARIO"]]
        # Incluir CNPJs que aparecem nos emplacamentos mas não na carteira
        cnpjs_emp_nao_cart = set(df_emp["CNPJ_NORM"].unique()) - set(df_cart["CNPJ_NORM"].unique())
        if cnpjs_emp_nao_cart:
            df_extras = nomes_emp.loc[nomes_emp.index.isin(cnpjs_emp_nao_cart)].reset_index()
            df_extras = df_extras.rename(columns={"NOMEPROPRIETARIO":"Nome","CPFCNPJPROPRIETARIO":"CPF/CNPJ"})
            df_extras["VENDEDOR"] = df_extras["CNPJ_NORM"].apply(
                lambda c: get_consultor_distribuido(str(c), vendedores_ativos)
            )
            df_dist = pd.concat([df_dist, df_extras], ignore_index=True)

    if perfil in ("gestor","gerente"):
        vends = ["Todos"] + sorted(df_dist["VENDEDOR"].dropna().unique().tolist())
        sel_vend = st.selectbox("Vendedor:", vends)
        cart_view = df_dist.copy() if sel_vend == "Todos" else df_dist[df_dist["VENDEDOR"] == sel_vend].copy()
    else:
        cart_view = df_dist[norm_str_series(df_dist["VENDEDOR"]) == norm_str(cons_key)].copy()

    total_cart = len(cart_view)

    if df_emp is not None:
        ultima_emp = df_emp.groupby("CNPJ_NORM")["Data emplacamento"].max().reset_index()
        cart_view = cart_view.merge(ultima_emp, on="CNPJ_NORM", how="left")
        def calc_meses(x):
            try:
                if pd.isna(x): return 999
                rd = relativedelta(today, pd.Timestamp(x))
                return int(rd.years*12 + rd.months)
            except: return 999
        cart_view["MesesSem"] = cart_view["Data emplacamento"].apply(calc_meses)
    else:
        cart_view["Data emplacamento"] = pd.NaT
        cart_view["MesesSem"] = 999

    cart_view["MesesSem"] = pd.to_numeric(cart_view["MesesSem"], errors="coerce").fillna(999).astype(int)
    inativos = cart_view[cart_view["MesesSem"] > 48]

    k1,k2,k3,k4 = st.columns(4)
    with k1: st.markdown(f'<div class="kpi-card"><div class="kpi-label">Total Carteira</div><div class="kpi-value">{total_cart}</div></div>', unsafe_allow_html=True)
    with k2: st.markdown(f'<div class="kpi-card"><div class="kpi-label">Inativos +4 anos</div><div class="kpi-value red">{len(inativos)}</div></div>', unsafe_allow_html=True)
    with k3: st.markdown(f'<div class="kpi-card"><div class="kpi-label">Ativos ≤4 anos</div><div class="kpi-value green">{total_cart-len(inativos)}</div></div>', unsafe_allow_html=True)
    with k4:
        if df_emp is not None:
            tot_e = len(df_emp[df_emp["CNPJ_NORM"].isin(cart_view["CNPJ_NORM"])])
            st.markdown(f'<div class="kpi-card"><div class="kpi-label">Emplacamentos</div><div class="kpi-value blue">{tot_e}</div></div>', unsafe_allow_html=True)

    st.markdown('<div class="sec-title">🔴 Para Revisão (inativos +4 anos)</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="alert-red">⚠️ <strong>{len(inativos)} clientes</strong> sem comprar há mais de 4 anos</div>', unsafe_allow_html=True)
    if not inativos.empty:
        cols_in = [c for c in ["Nome","CPF/CNPJ","VENDEDOR","Data emplacamento","MesesSem"] if c in inativos.columns]
        in_show = inativos[cols_in].copy()
        if "Data emplacamento" in in_show.columns:
            in_show["Data emplacamento"] = pd.to_datetime(in_show["Data emplacamento"], errors="coerce").dt.strftime("%d/%m/%Y").fillna("Nunca")
        in_show["MesesSem"] = in_show["MesesSem"].apply(lambda x: f"{x} meses" if x<999 else "Sem registro")
        in_show = in_show.rename(columns={"Data emplacamento":"Última Compra","MesesSem":"Meses Sem"})
        st.dataframe(in_show, use_container_width=True, hide_index=True)
        buf = BytesIO(); inativos.to_excel(buf, index=False, engine="openpyxl"); buf.seek(0)
        st.download_button("📥 Exportar Inativos", buf, file_name="carteira_revisao.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.markdown('<div class="sec-title">🏆 Maiores Compradores</div>', unsafe_allow_html=True)
    if df_emp is not None and not cart_view.empty:
        emp_c = df_emp[df_emp["CNPJ_NORM"].isin(cart_view["CNPJ_NORM"])].copy()
        top_c = emp_c.groupby(["CNPJ_NORM","NOMEPROPRIETARIO","NO_CIDADE"]).agg(
            Total=("Chassi","count"),
            Nigris=("Concessionário", lambda x: is_denigris(x).sum()),
            Ultima=("Data emplacamento","max"),
            Marca=("Marca", lambda x: x.mode()[0] if not x.empty else "—"),
        ).reset_index().sort_values("Total", ascending=False).head(20)
        top_c["% Nigris"] = (top_c["Nigris"]/top_c["Total"]*100).round(0).astype(int).astype(str)+"%"
        top_c["Ultima"] = pd.to_datetime(top_c["Ultima"]).dt.strftime("%d/%m/%Y")
        top_c = top_c.drop(columns=["CNPJ_NORM"]).rename(columns={"NOMEPROPRIETARIO":"Nome","NO_CIDADE":"Cidade","Total":"Total","Ultima":"Última Compra"})
        top_c.insert(0,"#", range(1,len(top_c)+1))
        st.dataframe(top_c, use_container_width=True, hide_index=True)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PÁGINA: PAINEL GERAL
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
elif pagina == "painel":
    st.markdown("""<div class="page-header"><h1>📊 Painel Geral</h1>
    <p>Visão consolidada do mercado</p></div>""", unsafe_allow_html=True)

    if perfil != "gestor":
        st.warning("⚠️ Acesso restrito ao administrador."); st.stop()
    if df_emp is None:
        st.warning("⚠️ Dados não carregados."); st.stop()

    hoje_p = pd.Timestamp.now()
    anos_d = sorted([int(a) for a in df_emp["Ano"].dropna().unique() if int(a) <= hoje_p.year], reverse=True)

    fp1, fp2, fp3 = st.columns(3)
    with fp1: sel_ano_p = st.selectbox("Ano:", anos_d, index=0, key="p_ano")
    with fp2:
        meses_ano_p = sorted(df_emp[df_emp["Ano"]==sel_ano_p]["Mes"].dropna().astype(int).unique().tolist())
        ops_mes_p = ["Todos os meses"] + [MESES_PT[m] for m in meses_ano_p]
        sel_mes_p = st.selectbox("Mês:", ops_mes_p, index=0, key="p_mes")
    with fp3:
        sel_tipo_p = st.selectbox("Tipo:", ["Todos","Somente Carteira","Fora da Carteira"], key="p_tipo")

    df_p = df_emp[df_emp["Ano"] == sel_ano_p].copy()
    if sel_mes_p != "Todos os meses":
        sel_mes_p_num = [k for k,v in MESES_PT.items() if v == sel_mes_p][0]
        df_p = df_p[df_p["Mes"] == sel_mes_p_num]
    periodo_lbl = f"{sel_mes_p}/{sel_ano_p}" if sel_mes_p != "Todos os meses" else str(sel_ano_p)

    todos_cnpjs_cart_p = set(df_cart["CNPJ_NORM"].dropna().unique()) if df_cart is not None else set()
    if sel_tipo_p == "Somente Carteira": df_p = df_p[df_p["CNPJ_NORM"].isin(todos_cnpjs_cart_p)]
    elif sel_tipo_p == "Fora da Carteira": df_p = df_p[~df_p["CNPJ_NORM"].isin(todos_cnpjs_cart_p)]

    if df_p.empty: st.info(f"Sem dados para {periodo_lbl}."); st.stop()

    total_p = len(df_p); nigris_p = int(is_denigris(df_p["Concessionário"]).sum())
    conc_p = total_p - nigris_p; ms_p = round(nigris_p/total_p*100,1) if total_p else 0
    clientes_p = df_p["CNPJ_NORM"].nunique()
    sem_cart_p = df_p[~df_p["CNPJ_NORM"].isin(todos_cnpjs_cart_p)]["CNPJ_NORM"].nunique()
    venda_perd = df_p[df_p["CNPJ_NORM"].isin(todos_cnpjs_cart_p) & ~is_denigris(df_p["Concessionário"])]["CNPJ_NORM"].nunique()

    k1,k2,k3,k4,k5,k6 = st.columns(6)
    for col, lbl, val, cls in [(k1,"Emplacamentos",f"{total_p:,}",""),
                                (k2,"Clientes",f"{clientes_p:,}",""),
                                (k3,"De Nigris",str(nigris_p),"green"),
                                (k4,"Concorrência",str(conc_p),"red"),
                                (k5,"Market Share",f"{ms_p}%","blue"),
                                (k6,"Vendas Perdidas",str(venda_perd),"red")]:
        with col: st.markdown(f'<div class="kpi-card"><div class="kpi-label">{lbl}</div><div class="kpi-value {cls}">{val}</div></div>', unsafe_allow_html=True)

    st.markdown('<div class="sec-title">🔴 Principal Concorrência</div>', unsafe_allow_html=True)
    df_p_conc = df_p[~is_denigris(df_p["Concessionário"])]
    conc_rank = df_p_conc.groupby("Concessionário").agg(Emplacamentos=("Chassi","count"),Clientes=("CNPJ_NORM","nunique")).reset_index().sort_values("Emplacamentos",ascending=False)
    conc_rank["% Total"] = (conc_rank["Emplacamentos"]/total_p*100).round(1).astype(str)+"%"
    if not conc_rank.empty:
        ca, cb = st.columns([3,2])
        with ca:
            t10 = conc_rank.head(10).copy(); t10["Nome_curto"] = t10["Concessionário"].str[:35]
            cores_c = ["#c8a84b" if i==0 else "#0a1628" for i in range(len(t10))]
            fig_c = go.Figure(go.Bar(x=t10["Emplacamentos"],y=t10["Nome_curto"],orientation="h",
                marker_color=cores_c,text=t10["Emplacamentos"],textposition="outside"))
            fig_c.update_layout(plot_bgcolor="#fff",paper_bgcolor="#fff",font_color="#4a5568",height=300,
                xaxis=dict(showgrid=True,gridcolor="#f0f2f7"),yaxis=dict(showgrid=False,autorange="reversed"),
                margin=dict(t=20,b=10,l=10,r=60),title=f"Top Concorrentes — {periodo_lbl}")
            st.plotly_chart(fig_c, use_container_width=True)
        with cb:
            cshow = conc_rank.head(10).copy(); cshow.insert(0,"Pos.",range(1,len(cshow)+1))
            cshow["Concessionário"] = cshow["Concessionário"].str[:38]
            st.dataframe(cshow[["Pos.","Concessionário","Emplacamentos","Clientes","% Total"]], use_container_width=True, hide_index=True)

    st.markdown('<div class="sec-title">⭐ Maiores Clientes Sem Cadastro</div>', unsafe_allow_html=True)
    df_sem = df_p[~df_p["CNPJ_NORM"].isin(todos_cnpjs_cart_p)]
    if not df_sem.empty:
        top_sem = df_sem.groupby(["CNPJ_NORM","NOMEPROPRIETARIO","NO_CIDADE"]).agg(
            Emplacamentos=("Chassi","count"), Nigris=("Concessionário",lambda x: is_denigris(x).sum()),
            Ultima=("Data emplacamento","max"), MarcaFreq=("Marca",lambda x: x.mode()[0] if not x.empty else "—"),
        ).reset_index().sort_values("Emplacamentos",ascending=False).head(20)
        top_sem["% Nigris"] = (top_sem["Nigris"]/top_sem["Emplacamentos"]*100).round(0).astype(int).astype(str)+"%"
        top_sem["Ultima"] = pd.to_datetime(top_sem["Ultima"]).dt.strftime("%d/%m/%Y")
        top_sem.insert(0,"#",range(1,len(top_sem)+1))
        st.markdown(f'<div class="alert-blue">💡 <strong>{sem_cart_p} clientes</strong> emplacaram sem estar na carteira — potencial de prospecção.</div>', unsafe_allow_html=True)
        st.dataframe(top_sem[["#","NOMEPROPRIETARIO","NO_CIDADE","Emplacamentos","MarcaFreq","% Nigris","Ultima"]].rename(columns={"NOMEPROPRIETARIO":"Cliente","NO_CIDADE":"Cidade","MarcaFreq":"Marca","Ultima":"Última Compra"}), use_container_width=True, hide_index=True)

    # Todos os emplacamentos
    mostrar_todos_p = st.toggle(f"📋 Ver todos os emplacamentos ({len(df_p):,})", value=False)
    if mostrar_todos_p:
        _cols_p = [c for c in ["Data emplacamento","CPFCNPJPROPRIETARIO","NOMEPROPRIETARIO","Modelo","Marca","Concessionário","NO_CIDADE"] if c in df_p.columns]
        det_p = df_p[_cols_p].copy()
        det_p["Data emplacamento"] = pd.to_datetime(det_p["Data emplacamento"], errors="coerce").dt.strftime("%d/%m/%Y")
        det_p["De Nigris"] = is_denigris(df_p["Concessionário"]).map({True:"✅",False:"—"})
        det_p["Na Carteira"] = df_p["CNPJ_NORM"].isin(todos_cnpjs_cart_p).map({True:"✅",False:"—"})
        det_p = det_p.rename(columns={"Data emplacamento":"Data","CPFCNPJPROPRIETARIO":"CNPJ","NOMEPROPRIETARIO":"Cliente","Concessionário":"Concessionária","NO_CIDADE":"Cidade"})
        st.dataframe(det_p.sort_values("Data"), use_container_width=True, hide_index=True)
        buf_p = BytesIO(); det_p.to_excel(buf_p, index=False, engine="openpyxl"); buf_p.seek(0)
        st.download_button(f"📥 Exportar todos", buf_p, file_name=f"emp_admin_{periodo_lbl.replace('/','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PÁGINA: GESTÃO
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
elif pagina == "gestao":
    st.markdown("""<div class="page-header"><h1>📈 Gestão & Performance</h1>
    <p>Análise por período e consultor</p></div>""", unsafe_allow_html=True)
    if df_emp is None: st.warning("⚠️ Dados não carregados."); st.stop()

    anos_g = sorted([int(a) for a in df_emp["Ano"].dropna().unique()], reverse=True)
    vends_g = ["Todos"] + get_vendedores_ativos()
    meses_dados = sorted(df_emp["Mes"].unique().tolist())
    mes_lbls_dados = [MESES_PT[m] for m in meses_dados]

    f1,f2,f3 = st.columns(3)
    with f1: anos_s = st.multiselect("Anos", anos_g, default=anos_g)
    with f2:
        mes_s = st.multiselect("Meses", mes_lbls_dados, default=mes_lbls_dados)
        mes_nums = [k for k,v in MESES_PT.items() if v in mes_s]
    with f3: cons_s = st.selectbox("Consultor", vends_g)

    df_g = df_emp.copy()
    if anos_s: df_g = df_g[df_g["Ano"].isin(anos_s)]
    if mes_nums: df_g = df_g[df_g["Mes"].isin(mes_nums)]

    vends_atv_g = get_vendedores_ativos()
    if df_cart is not None:
        cart_g = df_cart[["CNPJ_NORM","VENDEDOR"]].drop_duplicates(subset=["CNPJ_NORM"]).rename(columns={"VENDEDOR":"_VG"})
        df_g = df_g.merge(cart_g, on="CNPJ_NORM", how="left")
        def _vg(row):
            v = row.get("_VG","")
            if pd.notna(v) and str(v).strip() not in ["—","nan","None",""]:
                return str(v).strip()
            return get_consultor_distribuido(str(row.get("CNPJ_NORM","")), vends_atv_g)
        df_g["Consultor"] = df_g.apply(_vg, axis=1)
    else:
        df_g["Consultor"] = df_g["CNPJ_NORM"].apply(lambda c: get_consultor_distribuido(str(c), vends_atv_g))

    if cons_s != "Todos":
        df_g = df_g[df_g["Consultor"] == cons_s]

    if df_g.empty: st.warning("Sem dados."); st.stop()

    total_g = len(df_g); nigris_g = int(is_denigris(df_g["Concessionário"]).sum())
    k1,k2,k3,k4,k5 = st.columns(5)
    for c, lbl, v, cls in [(k1,"Emplacamentos",f"{total_g:,}",""),(k2,"Clientes",df_g["CNPJ_NORM"].nunique(),""),
                            (k3,"De Nigris",nigris_g,"green"),(k4,"Concorrente",total_g-nigris_g,"red"),
                            (k5,"Market Share",f"{round(nigris_g/total_g*100,1) if total_g else 0}%","blue")]:
        with c: st.markdown(f'<div class="kpi-card"><div class="kpi-label">{lbl}</div><div class="kpi-value {cls}">{v}</div></div>', unsafe_allow_html=True)

    st.markdown('<div class="sec-title">👥 Por Consultor</div>', unsafe_allow_html=True)
    perf = df_g.groupby("Consultor").agg(
        Total=("Chassi","count"), Nigris=("Concessionário",lambda x: is_denigris(x).sum()),
        Conc=("Concessionário",lambda x: (~is_denigris(x)).sum()), Clientes=("CNPJ_NORM","nunique"),
    ).reset_index().sort_values("Total",ascending=False)
    perf["% Nigris"] = (perf["Nigris"]/perf["Total"]*100).round(1).astype(str)+"%"
    fig_cv = go.Figure()
    fig_cv.add_trace(go.Bar(name="De Nigris",x=perf["Consultor"],y=perf["Nigris"],marker_color="#0a1628"))
    fig_cv.add_trace(go.Bar(name="Concorrente",x=perf["Consultor"],y=perf["Conc"],marker_color="#c8a84b"))
    fig_cv.update_layout(barmode="stack",plot_bgcolor="#fff",paper_bgcolor="#fff",font_color="#4a5568",height=280,
        xaxis=dict(showgrid=False,tickangle=-25),yaxis=dict(showgrid=True,gridcolor="#f0f2f7"),
        legend=dict(orientation="h",y=-0.3),margin=dict(t=10,b=60,l=10,r=10))
    st.plotly_chart(fig_cv, use_container_width=True)
    st.dataframe(perf.rename(columns={"Consultor":"Consultor","Total":"Total","Nigris":"De Nigris","Conc":"Concorrente","% Nigris":"% De Nigris"}), use_container_width=True, hide_index=True)

    st.markdown('<div class="sec-title">📅 Evolução Mensal</div>', unsafe_allow_html=True)
    df_g["AnoMes"] = df_g["Ano"].astype(str)+"-"+df_g["Mes"].astype(str).str.zfill(2)
    evol = df_g.groupby("AnoMes").agg(Total=("Chassi","count"),Nigris=("Concessionário",lambda x: is_denigris(x).sum())).reset_index().sort_values("AnoMes")
    evol["Conc"] = evol["Total"]-evol["Nigris"]
    fig_ev = go.Figure()
    fig_ev.add_trace(go.Scatter(x=evol["AnoMes"],y=evol["Nigris"],name="De Nigris",mode="lines+markers",
        line=dict(color="#0a1628",width=2.5),fill="tozeroy",fillcolor="rgba(10,22,40,0.06)"))
    fig_ev.add_trace(go.Scatter(x=evol["AnoMes"],y=evol["Conc"],name="Concorrente",mode="lines+markers",
        line=dict(color="#c8a84b",width=2,dash="dot")))
    fig_ev.update_layout(plot_bgcolor="#fff",paper_bgcolor="#fff",font_color="#4a5568",height=260,
        legend=dict(orientation="h",y=-0.35),xaxis=dict(showgrid=False,tickangle=-40),
        yaxis=dict(showgrid=True,gridcolor="#f0f2f7"),margin=dict(t=10,b=70,l=10,r=10))
    st.plotly_chart(fig_ev, use_container_width=True)

    ex1, ex2 = st.columns(2)
    with ex1:
        _exp_c = [c for c in ["Data emplacamento","NOMEPROPRIETARIO","NO_CIDADE","Marca","Modelo","Concessionário"] if c in df_g.columns]
        buf=BytesIO(); df_g[_exp_c].to_excel(buf,index=False,engine="openpyxl"); buf.seek(0)
        st.download_button("📄 Base Filtrada", buf, file_name="base_filtrada.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with ex2:
        buf2=BytesIO(); perf.to_excel(buf2,index=False,engine="openpyxl"); buf2.seek(0)
        st.download_button("👥 Consultores", buf2, file_name="consultores.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PÁGINA: OPORTUNIDADES — corrigida, sem erros vermelhos
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
elif pagina == "oportunidades":
    st.markdown("""<div class="page-header"><h1>🎯 Oportunidades</h1>
    <p>Prospects quentes · Inativos · Concorrência</p></div>""", unsafe_allow_html=True)

    if df_emp is None: st.warning("⚠️ Dados não carregados."); st.stop()

    today = pd.Timestamp.now()
    vendedores_ativos_o = get_vendedores_ativos()

    # Filtrar base por perfil
    df_opp = df_emp.copy()
    if perfil == "vendedor":
        if df_cart is not None:
            cnpjs_cart_v = set(df_cart[norm_str_series(df_cart["VENDEDOR"]) == norm_str(cons_key)]["CNPJ_NORM"].dropna())
        else:
            cnpjs_cart_v = set()
        todos_cnpjs_o = set(df_emp["CNPJ_NORM"].unique())
        cnpjs_dist_o  = {c for c in todos_cnpjs_o if c not in cnpjs_cart_v and
                         get_consultor_distribuido(str(c), vendedores_ativos_o) == cons_key}
        meus_cnpjs_o  = cnpjs_cart_v | cnpjs_dist_o
        df_opp = df_emp[df_emp["CNPJ_NORM"].isin(meus_cnpjs_o)].copy()
        st.markdown(f'<div class="alert-blue">📍 Exibindo sua carteira: <strong>{cons_key.title()}</strong></div>', unsafe_allow_html=True)

    tab1, tab2, tab3 = st.tabs(["🔴 Inativos +12m","🔥 Próxima Compra","⚠️ Concorrente"])

    # ── TAB 1: Inativos ──
    with tab1:
        try:
            ul = df_opp.groupby("CNPJ_NORM")["Data emplacamento"].max().reset_index()
            tc2 = df_opp.groupby("CNPJ_NORM").size().reset_index(name="TotalCompras")
            inf = df_opp.groupby("CNPJ_NORM").agg(Nome=("NOMEPROPRIETARIO","first"),CNPJ=("CPFCNPJPROPRIETARIO","first"),Cidade=("NO_CIDADE","first")).reset_index()
            di = ul.merge(tc2,on="CNPJ_NORM").merge(inf,on="CNPJ_NORM")
            di = di[di["Data emplacamento"].notna()].copy()
            di["Data emplacamento"] = pd.to_datetime(di["Data emplacamento"], errors="coerce")
            di = di[di["Data emplacamento"].notna()].copy()

            def _meses_desde(x):
                try:
                    rd = relativedelta(today, pd.Timestamp(x))
                    return int(rd.years*12 + rd.months)
                except Exception:
                    return 0

            di["Meses"] = di["Data emplacamento"].apply(_meses_desde)
            di["Meses"] = pd.to_numeric(di["Meses"], errors="coerce").fillna(0).astype(int)
            di = di[di["Meses"] > 12].sort_values("Meses", ascending=False)
            if df_cart is not None:
                vm = df_cart.set_index("CNPJ_NORM")["VENDEDOR"].to_dict()
                di["Vendedor"] = di["CNPJ_NORM"].map(vm).fillna("—")
            st.markdown(f'<div class="alert-red">🚨 <strong>{len(di)} clientes</strong> há mais de 12 meses sem comprar</div>', unsafe_allow_html=True)
            _dic = [c for c in ["Nome","CNPJ","Cidade","Data emplacamento","Meses","TotalCompras"] if c in di.columns]
            di_s = di[_dic].copy()
            di_s["Data emplacamento"] = di_s["Data emplacamento"].dt.strftime("%d/%m/%Y")
            di_s = di_s.rename(columns={"Data emplacamento":"Última Compra","Meses":"Meses Sem","TotalCompras":"Total Compras"})
            st.dataframe(di_s, use_container_width=True, hide_index=True)
            buf=BytesIO(); di_s.to_excel(buf,index=False,engine="openpyxl"); buf.seek(0)
            st.download_button("📥 Exportar Inativos", buf, file_name="inativos.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e_tab1:
            st.markdown(f'<div class="alert-gray">ℹ️ Não foi possível calcular inativos: {str(e_tab1)[:120]}</div>', unsafe_allow_html=True)

    # ── TAB 2: Próxima Compra ──
    with tab2:
        quentes = []
        erros_pred = 0
        for cnpj_q, grp_q in df_opp.groupby("CNPJ_NORM"):
            try:
                dts_q = grp_q["Data emplacamento"].dropna().tolist()
                if len(dts_q) < 2:
                    continue
                pl, pd_ = calc_prediction(dts_q)
                if pd_ is None:
                    continue
                try:
                    dm = relativedelta(pd_, today)
                    m_q = dm.years * 12 + dm.months
                except Exception:
                    continue
                if -1 <= m_q <= 3:
                    r_q = grp_q.sort_values("Data emplacamento", ascending=False).iloc[0]
                    quentes.append({
                        "Nome": str(r_q.get("NOMEPROPRIETARIO",""))[:40],
                        "CNPJ": str(r_q.get("CPFCNPJPROPRIETARIO","")),
                        "Cidade": str(r_q.get("NO_CIDADE","")),
                        "Previsão": pl,
                        "Meses": m_q,
                        "Total": len(grp_q)
                    })
            except Exception:
                erros_pred += 1
                continue

        if quentes:
            dq = pd.DataFrame(quentes).sort_values("Meses")
            st.markdown(f'<div class="alert-green">🔥 <strong>{len(dq)} clientes</strong> com previsão de compra nos próximos 90 dias</div>', unsafe_allow_html=True)
            if erros_pred > 0:
                st.markdown(f'<div class="alert-gray">ℹ️ {erros_pred} registros ignorados por datas inválidas.</div>', unsafe_allow_html=True)
            st.dataframe(dq, use_container_width=True, hide_index=True)
            buf=BytesIO(); dq.to_excel(buf,index=False,engine="openpyxl"); buf.seek(0)
            st.download_button("📥 Exportar Prospects", buf, file_name="prospects.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.info("Nenhum prospect quente identificado no momento.")
            if erros_pred > 0:
                st.markdown(f'<div class="alert-gray">ℹ️ {erros_pred} registros ignorados por datas inválidas.</div>', unsafe_allow_html=True)

    # ── TAB 3: Concorrência ──
    with tab3:
        try:
            conc_df = df_opp[~is_denigris(df_opp["Concessionário"])].copy()
            cs = conc_df.groupby(["CNPJ_NORM","NOMEPROPRIETARIO","NO_CIDADE"]).agg(
                Qtd=("Chassi","count"), UltimaConc=("Data emplacamento","max"),
                PrincipalConc=("Concessionário",lambda x: x.mode()[0] if not x.empty else "—")
            ).reset_index().sort_values("Qtd",ascending=False)
            cs["UltimaConc"] = pd.to_datetime(cs["UltimaConc"], errors="coerce").dt.strftime("%d/%m/%Y")
            cs = cs.drop(columns=["CNPJ_NORM"]).rename(columns={"NOMEPROPRIETARIO":"Cliente","NO_CIDADE":"Cidade","Qtd":"Qtd Concorrente","UltimaConc":"Última Compra Conc.","PrincipalConc":"Principal Concorrente"})
            st.markdown(f'<div class="alert-yellow">⚠️ <strong>{len(cs)} clientes</strong> emplacaram em concorrentes</div>', unsafe_allow_html=True)
            st.dataframe(cs, use_container_width=True, hide_index=True)
            buf=BytesIO(); cs.to_excel(buf,index=False,engine="openpyxl"); buf.seek(0)
            st.download_button("📥 Exportar Concorrentes", buf, file_name="concorrentes.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e_tab3:
            st.markdown(f'<div class="alert-gray">ℹ️ Erro ao processar dados de concorrência: {str(e_tab3)[:120]}</div>', unsafe_allow_html=True)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PÁGINA: ADMIN
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
elif pagina == "admin":
    if perfil != "gestor":
        st.error("Acesso restrito ao Administrador."); st.stop()

    st.markdown("""<div class="page-header"><h1>⚙️ Administração</h1>
    <p>Usuários · Dados · Configurações</p></div>""", unsafe_allow_html=True)

    tab_u, tab_d = st.tabs(["👥 Usuários","📂 Gerenciar Dados"])

    with tab_u:
        USERS = st.session_state.users_db
        token_gh, repo_gh, branch_gh = _gh_secrets()
        if token_gh and repo_gh:
            api_test = f"https://api.github.com/repos/{repo_gh}/contents/data/users.json"
            content_test, _ = _gh_get_file(api_test, token_gh)
            if content_test is not None:
                st.markdown(f'<div class="alert-blue">✅ <strong>GitHub conectado</strong> — repo: <code>{repo_gh}</code> · branch: <code>{branch_gh}</code></div>', unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="alert-yellow">⚠️ GitHub com problema de conexão.</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="alert-red">🔴 GitHub não configurado. Usuários não persistem entre reboots.</div>', unsafe_allow_html=True)

        st.markdown('<div class="sec-title">👥 Usuários Cadastrados</div>', unsafe_allow_html=True)
        rows = []
        for login, ud in USERS.items():
            ua = ud.get("ultimo_acesso")
            ua_fmt = "Nunca"
            if ua:
                try:
                    ua_dt = datetime.datetime.fromisoformat(str(ua)[:19])
                    ua_fmt = ua_dt.strftime("%d/%m/%Y %H:%M") + " (BRT)"
                except Exception:
                    ua_fmt = str(ua)[:16]
            rows.append({"Login":login,"Nome":ud.get("nome",""),"Perfil":ud.get("perfil",""),"Último Acesso":ua_fmt})
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        st.markdown('<div class="sec-title">➕ Criar / Editar Usuário</div>', unsafe_allow_html=True)
        st.markdown('<div class="alert-blue">💡 Para <strong>vendedores</strong>: selecione o nome exato como aparece na planilha Carteira. Clientes não cadastrados são distribuídos automaticamente entre todos os vendedores ativos.</div>', unsafe_allow_html=True)

        eu1, eu2 = st.columns(2)
        with eu1:
            new_login  = st.text_input("Login (maiúsculas)", placeholder="EX: RENATA").strip().upper()
            new_nome   = st.text_input("Nome completo", placeholder="Renata Bellon")
            new_perfil = st.selectbox("Perfil", ["vendedor","gerente","gestor"])
            if new_perfil == "vendedor" and df_cart is not None:
                vends_cart = sorted(df_cart["VENDEDOR"].dropna().unique().tolist())
                ops_c = ["— Selecione —"] + vends_cart
                sel_c = st.selectbox("Consultor vinculado:", ops_c)
                new_cons_key = "" if sel_c == "— Selecione —" else norm_str(sel_c)
            else:
                new_cons_key = ""
                if new_perfil == "vendedor":
                    new_cons_key = st.text_input("Consultor (manual):", placeholder="NOME COMO NA CARTEIRA").strip().upper()

        with eu2:
            new_senha  = st.text_input("Senha", type="password")
            new_senha2 = st.text_input("Confirmar senha", type="password")

        if st.button("💾 Salvar Usuário", use_container_width=True):
            if not new_login or not new_nome or not new_senha:
                st.error("Preencha login, nome e senha.")
            elif new_senha != new_senha2:
                st.error("Senhas não conferem.")
            elif new_perfil == "vendedor" and not new_cons_key:
                st.error("Informe o consultor vinculado.")
            else:
                USERS[new_login] = {
                    "senha_hash": hash_senha(new_senha),
                    "perfil": new_perfil,
                    "nome": new_nome,
                    "consultor_key": new_cons_key if new_cons_key else norm_str(new_login),
                    "ultimo_acesso": USERS.get(new_login, {}).get("ultimo_acesso")
                }
                gh_ok, gh_err = save_users(USERS)
                st.session_state.users_db = USERS
                msg = f"✅ Usuário **{new_login}** salvo!"
                msg += " Sincronizado com GitHub." if gh_ok else f" (GitHub: {gh_err})"
                st.success(msg)
                st.rerun()

        st.markdown('<div class="sec-title">🗑️ Excluir Usuário</div>', unsafe_allow_html=True)
        outros = [l for l in USERS.keys() if l != u_key]
        if outros:
            del_login = st.selectbox("Selecionar para excluir:", outros)
            if st.button("❌ Excluir Usuário", use_container_width=True):
                del USERS[del_login]
                gh_ok, gh_err = save_users(USERS)
                st.session_state.users_db = USERS
                st.success(f"✅ Usuário {del_login} excluído.")
                st.rerun()
        else:
            st.info("Nenhum outro usuário para excluir.")

    with tab_d:
        st.markdown('<div class="sec-title">📊 Status dos Dados</div>', unsafe_allow_html=True)
        s1,s2,s3 = st.columns(3)
        with s1: st.markdown('<div class="kpi-card"><div class="kpi-label">Distribuição</div><div class="kpi-value" style="font-size:14px;">✅ Dinâmica</div></div>', unsafe_allow_html=True)
        with s2:
            sc = f"✅ {len(df_cart):,}" if df_cart is not None else "⚠️ Não carregada"
            st.markdown(f'<div class="kpi-card"><div class="kpi-label">Carteira</div><div class="kpi-value" style="font-size:14px;">{sc}</div></div>', unsafe_allow_html=True)
        with s3:
            if df_emp is not None:
                anos_e = sorted([int(a) for a in df_emp["Ano"].dropna().unique()])
                anos_str = f"{anos_e[0]}–{anos_e[-1]}" if len(anos_e)>1 else str(anos_e[0]) if anos_e else "?"
                se = f"✅ {len(df_emp):,}<br><small>{anos_str}</small>"
            else: se = "⚠️ Não carregado"
            st.markdown(f'<div class="kpi-card"><div class="kpi-label">Emplacamentos</div><div class="kpi-value" style="font-size:13px;">{se}</div></div>', unsafe_allow_html=True)

        st.markdown('<div class="alert-blue">💡 Substitua os arquivos no GitHub na pasta <code>data/</code>:<br>• <code>CARTEIRA_VANS.xlsx</code> · <code>EMPLACAMENTO APP VANS.xlsx</code><br>O sistema carrega automaticamente. Para repositório privado, configure <code>GH_TOKEN</code> nos Secrets.</div>', unsafe_allow_html=True)

        st.markdown('<div class="sec-title">📤 Upload Manual (temporário)</div>', unsafe_allow_html=True)
        col_u1, col_u2 = st.columns(2)
        with col_u1:
            st.markdown('<div class="upload-box"><div class="upload-title">📋 Carteira</div>', unsafe_allow_html=True)
            up_c = st.file_uploader("CARTEIRA.xlsx", type=["xlsx"], key="up_cart")
            if up_c:
                st.session_state.df_cart = load_carteira(BytesIO(up_c.getvalue()))
                st.session_state.dados_carregados = True
                st.success("✅ Carteira atualizada!")
            st.markdown('</div>', unsafe_allow_html=True)
        with col_u2:
            st.markdown('<div class="upload-box"><div class="upload-title">🚚 Emplacamentos</div>', unsafe_allow_html=True)
            up_e = st.file_uploader("Emplacamentos.xlsx", type=["xlsx"], key="up_emp", accept_multiple_files=True)
            if up_e:
                novos = 0
                for f in up_e:
                    if f.name not in st.session_state.emp_fontes:
                        st.session_state.df_emp_list.append(load_emplacamentos(BytesIO(f.getvalue()), label=f.name))
                        st.session_state.emp_fontes.append(f.name)
                        novos += 1
                if novos:
                    st.session_state["_df_emp_merged"] = merge_emp(st.session_state.df_emp_list)
                    st.session_state["_emp_merged_len"] = len(st.session_state.df_emp_list)
                    st.session_state.dados_carregados = True
                    st.success(f"✅ {novos} arquivo(s) adicionado(s)!")
            if st.session_state.emp_fontes:
                st.markdown("**Fontes:** " + ", ".join([f"`{f}`" for f in st.session_state.emp_fontes]))
                if st.button("🗑️ Limpar emplacamentos"):
                    st.session_state.df_emp_list=[]; st.session_state.emp_fontes=[]
                    st.session_state.pop("_df_emp_merged", None); st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

st.markdown('</div>', unsafe_allow_html=True)
